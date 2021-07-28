#!/usr/bin/python3
encoding="UTF-8"

#report table subscript
##############################################################
###Dependencies:
import sys
from sys import argv
import subprocess
import calc_ACMG
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, Fill, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from collections import OrderedDict
import time
import get_categories_fantom
import read_snps_new
import pandas as pd
from  pybiomart import Server
import panelapp
import bisect
import marrvel_input
import cProfile, pstats, io
import pickle
#############SCRIPTS USED FOR SMALL STEPS TROUGHOUT THE SCRIPT###############################

def parse_oe(oe, ell):
	"""reads the oe file or the HI/triplo file or the genehancer file
	, and returns the oe value for the
	gene indicated by the variable ell."""
	f=open(oe)
	f.readline()
	nn="nd"
	for el in f:
		line=el.split("\t")
		if line[0]==ell:
				nn=line[-1].strip()#returns the oe with the confidence interval, in all cases
				break
	f.close()
	return nn

def parse_acmg(infile):
	f=open(infile)
	dd=set()
	for i in f:
		line=i.split("\t")
		dd.add(line[2])
	f.close()
	return dd

def read_cytoband(infile, chrr, bp):
	"""Read the cytoband file and the breakpoint,
	and returns the cytoband correspondig to the breakpoint.
	Only the band is returned, not the cromosome."""
	f=open(infile)
	for i in f:
		line=i.split()
		if line[0]=="chr"+chrr:#the input file comes as "chrX", not just "X"
			if int(float(bp))>int(line[1]) and int(float(bp))<int(line[2]):
				aa=line[-1]
				break
	f.close()
	return aa

def merge_two_ordered_dicts(x,y):
	"""merges two ordered dict, with the care of adding
	the breakpoint_B on the right locale. Returns the
	merged dict"""
	z=OrderedDict()
	kh=list(y.keys())
	bpp=kh.index("Breakpoint_B")
	aa=0
	for key, value in x.items():
		if aa!=bpp+1:
			z[key]=value
			aa+=1
		else:
			z["Breakpoint_B"]=y["Breakpoint_B"]
			z[key]=value
			aa+=1
	return z

def read_chr_size(infile):
	"""reads the cromosome sizes to a dictionary as chrA:size. Returns
	the dictionary."""
	with open(infile) as fileobj:
		result = {row[0]: row[1] for line in fileobj for row in (line.split("\t"),)}
	return result

def check_genecard_existence(urll):
	"""Recieves a genecard link and confirms if it exists on the site.
		Returns true if exists, returns false if it do not exist"""
	bb=['Rscript reqs.r '+urll]#uses an auxiliary Rscript, that is runned by subprocess
	aa=subprocess.check_output(bb, shell=True, stderr=subprocess.STDOUT)
	if "FALSE" in str(aa):
		return "true"#returns as text not bolean
	else:
		return "false"

def makePQ(bpk):
	"""during the execution of the nomenclature, if there is no deletion
	nor duplication on the breakpoint, this method defines the breakpoint. If as
	breakpoint a region is given, it gets the middlepoint of the region. Otherwise
	returns the coordinate itself"""
	if "-" in bpk:
		gg=bpk.split("-")
		ss=int(gg[1])-int(gg[0])
		return put_comma(str(int(float(gg[0])+(ss/2))))
	else:
		return put_comma(str(int(float(bpk))))

def check_points(ws2):
	"""This function reads thoughout the already completed file and checks
	for more than one breakpoint, in case of deletion and duplication. This will allow the definition of
	the region of the table with backgroud. it returns the number of breakpoints defined on the table"""
	aa=0
	i=1
	while i<=ws2.max_row:
		n=ws2["A"+str(i)].value
		if n!=None:
			if n.startswith("der") or n.startswith("chr") or n.startswith("g.") or n.startswith("Breakpoint"):#looks for lines with the breakpoint paterns
				aa+=1
		i+=1
	return aa 

def read_lims(infile, chrr):
	"""reads the limits file for the chrr. returns a list of 
	ints with the begining of the first TAD, the end of the last
	TAD before the centromere, the begining of the first TAD after the
	centromere and the end of the last TAD."""
	f=open(infile)
	for i in f:
		if i.startswith(chrr+"\t") or i.startswith("chr"+chrr+"\t"):
			line=i.split("\t")
			break
	f.close()
	results = list(map(int, line[1:]))
	return results

def execute_hposim(terms, omimid):
        #bb=['Rscript phen_disease_for_tool.R  "HP:0011968" "HP:0001250" "HP:0012443" "HP:0000708" "HP:0000729" "OMIM:617616"']
        bb=['Rscript phen_disease_for_tool.R '+terms+' "OMIM:'+omimid.split('"')[-1]+'"']#ver como por a dar
        aa=subprocess.check_output(bb, shell=True, stderr=subprocess.STDOUT)
        #aa=subprocess.call(bb)#, stdout=subprocess.PIPE)
        gg=str(aa)
        ff=gg.split("[1]")[-1]
        return ff.split('"')[1]

#der(1) g.[chr8:55655297_qterinv::chr1:54303617_cen_qter]

def put_comma(v):#f'{value:,}
	if "-" in v:
		y=v.replace(",","")
		aa=y.split("-")
		return f'{int(float(aa[0])):,}'+"_"+f'{int(float(aa[1])):,}'
	elif "_" in v and "del" in v:
		y=v.replace(",","")
		u=y.replace("del","")
		aa=u.split("_")
		return f'{int(float(aa[0])):,}'+"_"+f'{int(float(aa[1])):,}'+"del"
	else:
		y=v.replace(",","")
		return f'{int(float(y)):,}'


def pq_nomenclature(chrA, chrB, bandA, bandB, pqAA, pqBB, isspec, isdelA, isdupA, isdelB, isdupB, typee, der, dic_chr_ref):
	"""makes the breakpoint nomenclature according to ISCN2020"""
	pqA=pqAA.strip()
	pqB=pqBB.strip()
	singA=pqA
	singB=pqB
	if typee=="ins":#chrA is the sink cromossome and chrB the source cromossome, in insertion
		if der==chrA:
			return "Breakpoint "+dic_chr_ref[chrA.replace("chr","")]+":g."+put_comma(pqA)+"ins["+dic_chr_ref[chrB.replace("chr","")]+":g."+put_comma(pqB)+"]"
		else:
			return "Breakpoint "+dic_chr_ref[chrB.replace("chr","")]+":g."+put_comma(pqB)+"del"
	elif isspec==True:
		return "Breakpoint "+chrA+bandA+": g.["+put_comma(pqAA)+"-"+put_comma(pqBB)+"]"
	aa=pqA.split("-")
	bb=pqB.split("-")
	if typee=="del":
		return "Breakpoint "+dic_chr_ref[chrA.replace("chr","")]+": g."+put_comma(pqA)+"_"+put_comma(pqB)+"del"
	elif typee=="dup":
		return "Breakpoint "+dic_chr_ref[chrA.replace("chr","")]+": g."+put_comma(pqA)+"_"+put_comma(pqB)+"dup"
	elif typee=="inv":
		return "Breakpoint "+dic_chr_ref[chrA.replace("chr","")]+":g.["+put_comma(pqA)+"_"+put_comma(pqB)+"inv]"
	elif typee=="trans":
		if der==chrA:
			if "p" in bandA and "p" in bandB:
				cc="Breakpoint der("+chrA+") "+dic_chr_ref[chrA.replace("chr","")]+":g.pter_"+put_comma(aa[0])+"delins["+dic_chr_ref[chrB.replace("chr","")]+":g.pter_"+put_comma(bb[-1])+"]"
			if "p" in bandA and "q" in bandB:
				cc="Breakpoint der("+chrA+") "+dic_chr_ref[chrA.replace("chr","")]+":g.pter_"+put_comma(aa[0])+"delins["+dic_chr_ref[chrB.replace("chr","")]+":g."+put_comma(bb[-1])+"_qterinv]"
			if "q" in bandA and "p" in bandB:
				cc="Breakpoint der("+chrA+") "+dic_chr_ref[chrA.replace("chr","")]+":g."+put_comma(aa[0])+"_qterdelins["+dic_chr_ref[chrB.replace("chr","")]+":g."+put_comma(bb[-1])+"_pterinv]"
			if "q" in bandA and "q" in bandB:
				cc="Breakpoint der("+chrA+") "+dic_chr_ref[chrA.replace("chr","")]+":g."+put_comma(aa[0])+"_qterdelins["+dic_chr_ref[chrB.replace("chr","")]+":g."+put_comma(bb[-1])+"_qter]"
		if der==chrB:
			if "p" in bandA and "p" in bandB:
				cc="Breakpoint der("+chrB+") "+dic_chr_ref[chrB.replace("chr","")]+":g.pter_"+put_comma(bb[0])+"delins["+dic_chr_ref[chrA.replace("chr","")]+":g.pter_"+put_comma(aa[-1])+"]"
			if "p" in bandA and "q" in bandB:
				cc="Breakpoint der("+chrB+") "+dic_chr_ref[chrB.replace("chr","")]+":g."+put_comma(bb[0])+"_qterdelins["+dic_chr_ref[chrA.replace("chr","")]+"g."+put_comma(aa[-1])+"_pterinv]"
			if "q" in bandA and "p" in bandB:
				cc="Breakpoint der("+chrB+") "+dic_chr_ref[chrB.replace("chr","")]+":g.pter_"+put_comma(bb[0])+"delins["+dic_chr_ref[chrA.replace("chr","")]+"g."+put_comma(aa[-1])+"_qterinv]"
			if "q" in bandA and "q" in bandB:
				cc="Breakpoint der("+chrB+") "+dic_chr_ref[chrB.replace("chr","")]+":g."+put_comma(bb[0])+"_qterdelins["+dic_chr_ref[chrA.replace("chr","")]+"g."+put_comma(aa[-1])+"_qter]"
		return cc
#########################################################TADS

def read_tads(tadfile, chrr, sz):
	"""reads the tads of chrr to a list. this list startswith 1
	and ends witgh sz. all elements of the list correspond to start and end
	of TADs. [1, start TAD1, end TAD1, start TAD2...., sz].Returns the list"""
	f=open(tadfile)
	l=[1]#starts with 1, the begining of the cromossome
	for i in f:
		if i.startswith(chrr+"\t") or i.startswith("chr"+chrr+"\t"):
			line=i.split("\t")
			l.append(int(line[1]))
			l.append(int(line[2]))
	f.close()
	if l[-1]!=int(sz):
		l.append(sz)#ends with the size of the cromosome, the end of it
	return l



def tads_by_reg(bp, tads_file, lims_file, params, v):
	"""when the user choses specific region instead of TADs, this function is activated. This function
	selects the TADs where the start and end of the specific region is located, and retrives a list
	with the TAD number relative to the breakpoint. The retrived list is then provided to the newtads
	function for normal processing"""
	print("lims_file", lims_file, v)
	print("params tads",  params["tads"])
	lims=read_lims(lims_file, params["tads"][v][0])#reads the limits of the telomeric and centromeric TADs to a list
	tads=read_tads(tads_file, params["tads"][v][0], lims[-1])#reads all the TADs of the respective cromossome to a list
	t1=bisect.bisect_left(tads, int(params["tads"][v][1]))#t is the breakpoint position at tads
	t2=bisect.bisect_left(tads, int(params["tads"][v][2]))#t is the breakpoint position at tads
	t3=t=bisect.bisect_left(tads, int(bp))
	if t1%2==1:
		t1-=1
	if t2%2==1:
		t2+=1
	if t3%2==1:
		return [int((t1-t3)), int((t2-t3))]
	else:
		return [int((t1-t3)/2), int((t2-t3)/2)]
		
		
def newtads(chrr, bp, tads_file, lims_file, params, v):
	"""using the limits and the list of TAD limits, it will retrive the
	TADs according to the users preference, and according to the breakpoint location.
	The function uses bisect to select the tads and determinate if the breakpoint is inside a TAD
	or not. Returns the dictionary tds with the referneced TADs, intertad regions, or genomic
	locationsas key (TAD-1, TBR, qter), and the respective limits as value """
	tds=OrderedDict()
	if chrr=="Y" or chrr=="chrY":
		l=params["tad"][0]
		st=bp-int(params["for_y"][l]/2)
		end=bp+int(params["for_y"][l]/2)
		if st<0:
			st=1
		if end> params["for_y"]["sz"]:
			end= params["for_y"]["sz"]
		tds["brTAD_"]=[st,end]
	elif params["ttype"]=="spec":
		to_run=tads_by_reg(bp, tads_file, lims_file, params, v)
	else:
		if len(params["tads"])==1:
			to_run=[params["tads"][0], params["tads"][0]]
		else:
			to_run=[params["tads"][0], params["tads"][1]]
	if chrr!="Y" and chrr!="chrY":
		lims=read_lims(lims_file, chrr)#reads the limits of the telomeric and centromeric TADs to a list
		cenl=read_lims(params["cen"], chrr)#centromeres
		tads=read_tads(tads_file, chrr, lims[-1])#reads all the TADs of the respective cromossome to a list
		cen=bisect.bisect_left(tads, lims[2])#cen is the position of the centromere relativelly to the tads list (tads)
		t=bisect.bisect_left(tads, int(bp))#t is the breakpoint position at tads
		if t%2==1 and bp==tads[t] and t!=len(tads)-1:
			t+=1
		if t%2==0:# if t is an 	even number, it means the breakpoint is inside a TAD
			for el in range(to_run[0], to_run[1]+1):
				if t+el!=cen and t+el+el-1>=1 and t+el+el-1<=len(tads)-1 and ((t+el+el<=cen and t<cen) or (t+el+el>=cen and t>cen)):
					if t+el==1:
						tds["pter_br_"]=[tads[0], tads[1]]
					elif t+el+el-1==len(tads)-1:
						tds["qter_br_"]=[tads[t+el+el-2], lims[-1]]	
					else:
						if el<0:
							tds["TAD"+str(el)+"_"]=[tads[t+el+el-1], tads[t+el+el]]
						elif el==0:
							tds["brTAD_"]=[tads[t-1], tads[t]]
						else:
							tds["TAD+"+str(el)+"_"]=[tads[t+el+el-1], tads[t+el+el]]
					if t+el+el+2<len(tads)-1 and tads[t+el+el+1]-tads[t+el+el]>1000 and el+1 <= to_run[1] and ((t+el+el+1<cen and t<cen) or (t+el+el+1>cen and t>cen)):
						tds["TBR_"+str(el)]=[tads[t+el+el], tads[t+el+el+1]]
				else:
					if bp>=lims[1] and bp<cenl[0]:
						tds["TBR_"]=[lims[1], cenl[0]]
					elif bp>=cenl[0] and bp<cenl[1]:
						tds["TBR_"]=[cenl[0], cenl[1]]
					elif bp>=cenl[1] and bp<lims[2]:
						tds["TBR_"]=[cenl[1], lims[2]]
		if t%2==1:# if it is an uneven number, it means the breakpoint is in an interTAD region
			for el in range(to_run[0], to_run[1]+1):
				if t+el!=cen and t+el>=1 and t+el+el-1<=len(tads)-1 and ((t+el+el-1<=cen and t<cen) or (t+el+el>=cen and t>cen)):
					if t+el==1:
						tds["pter_br_"]=[tads[0], tads[1]]
					elif t+el+el==len(tads)-1:
						tds["qter_br_"]=[tads[-1], lims[-1]]
					elif t+el!=cen:
						if el<0:
							tds["TAD"+str(el)+"_"]=[tads[t+el+el], tads[t+el+el+1]]
						elif el==0:
							tds["TBR_"]=[tads[t-1], tads[t]]
						else:
							tds["TAD+"+str(el)+"_"]	=[tads[t+el+el-2], tads[t+el+el-1]]
					if el>0 and tads[t+el+el]-tads[t+el+el-1]>1000 and el+1 <= to_run[1]  and el+1!=0 and t+el+el+1<len(tads)-1 and t+el!=1 and ((t+el+el<cen and t<cen) or (t+el+el>cen and t>cen)):
						tds["TBR_"+str(el)]=[tads[t+el+el-1], tads[t+el+el]]
					if el<0 and tads[t+el+el+2]-tads[t+el+el+1]>1000 and el+1 <= to_run[1]  and el+1!=0 and t+el+el+1<len(tads)-1 and t+el!=1 and ((t+el+el<=cen and t<cen) or (t+el+el>=cen and t>cen)):
						tds["TBR_"+str(el)]=[tads[t+el+el+1], tads[t+el+el+2]]
				else:
					if bp>=lims[1] and bp<cenl[0]:
						tds["TBR_"]=[lims[1], cenl[0]]
					elif bp>=cenl[0] and bp<cenl[1]:
						tds["TBR_"]=[cenl[0], cenl[1]]
					elif bp>=cenl[1] and bp<lims[2]:
						tds["TBR_"]=[cenl[1], lims[2]]
	return tds
	
def get_tdel(intnewt1, intnewt2):
	"""returns the delected region, when it is more than the brTAD. 
	Used by deals_with_deletions"""
	if "pter_br_" in intnewt1:
		co1=intnewt1["pter_br_"][1]
	if "qter_br_" in intnewt2:
		co2=intnewt2["qter_br_"][0]
	if "TBR_" in intnewt1:
		co1=intnewt1["TBR_"][1]
	if "brTAD_" in intnewt1: 
		co1=intnewt1["brTAD_"][1]
	if "TBR_" in intnewt2:
		co2=intnewt2["TBR_"][0]
	if "brTAD_" in intnewt2: 
		co2=intnewt2["brTAD_"][0]
	if int(co2)-int(co1)>100:
		return {"TDel":[co1,co2]}
	else:
		return {}

def deleted_tads( tdel, ch, params):
	"""Organizes the deleted tads defined in get_tdel on the TADs list, so it maintains the
	coordinate sort of them. Returns the params with the new tads and a list of the new tads to be analysed."""
	f=open(params["tad"][1])
	r=[]
	selected=OrderedDict()
	tads_inter=OrderedDict()
	numb=1
	for i in f:
		line=i.split("\t")
		if line[0]=="chr"+str(ch):
			if len(r)==0:
				r=[int(line[1]), int(line[2])]
				tads_inter["TDel_"+str(numb)]=[int(line[1]), int(line[2])]
				numb+=1
			else:
				if int(line[1])-r[1]>100:
					tads_inter["interTDel_"+str(numb)]=[int(r[1]), int(line[1])]
					numb+=1
					r=[int(line[1]), int(line[2])]
					tads_inter["TDel_"+str(numb)]=[int(line[1]), int(line[2])]
					numb+=1
				else:
					r=[int(line[1]), int(line[2])]
					tads_inter["TDel_"+str(numb)]=[int(line[1]), int(line[2])]
					numb+=1	
	f.close()
	for key, value in tads_inter.items():
		if int(tdel[0])<=value[0] and int(tdel[1])>=value[1]:
			selected[key]=value
			params["tads"].append(key)
	return selected, params
###################################Gene/Databases dependent part of the script################################################################
def search(filt, att, params):
	"""makes the search against the biomart database, using as filter arguments
	filt and retriving the atributes at att. Returns the response from the database"""
	if params["version"]=="hg19":
		server = Server(host='http://grch37.ensembl.org/')
		dataset = (server.marts['ENSEMBL_MART_ENSEMBL'].datasets['hsapiens_gene_ensembl'])
	elif params["version"]=="hg38":
		server = Server(host='http://www.ensembl.org')
		dataset = (server.marts['ENSEMBL_MART_ENSEMBL'].datasets['hsapiens_gene_ensembl'])
	try:
		response=dataset.query(attributes=att, filters=filt)
	except:
		response=pd.DataFrame(columns=['User_ID', 'UserName'])
	return response


def make_phen_nomenclature(dic, params):
	"""uses the pybiomart search, the OMIM inheritance file, the DDG2P file
	and de clingen file to create the phenotype fields of the script. The structure of
	the text depends on the data of the diferent databases. In the end returns a dictionary similar to
	the one inputed, with value[4] list completed with the phenotype. each element of value[4] corresponds
	to a diferent phenotype associated with the gene (key)."""
	entries=[]
	dicnew={}
	forsub=set()
	dd2p=get_categories_fantom.get_dd2p(params["DDG2P"])#reads the DDG2P file
	clingen=get_categories_fantom.clingen(params["clingen"])#reads the clingen file
	for key,value in dic.items():
		omim="ND"
		descr="ND"
		ddg="ND"
		cling="ND"
		if "NOOMIM" not in value[4]:#checks if the gene as an phenotype OMIM associated. 
			for el in value[4]:#goes trough the list that has the following structure: [[geneomimid, phenomimid1],[geneomimid, phenomimid2]...]
				inh=el[-1]
				forsub.add(el[-1].strip())
				omim=el[0]+"_"+inh+" "#prepares the format of the text for the omim part
				descr=el[1].split(";")[0]#add the description from omim
				if value[0] in dd2p:#checks if the gene is on DDG2P database
					dd2pp=dd2p[value[0]]#if it is
					if len(dd2pp[0])!=0 and el[1] in dd2pp[1]:
						ind=dd2pp[1].index(el[1])
						ddg=dd2p[value[0]][-1][0].replace("SUBSTITUTE",dd2pp[0][ind])
				if value[0] in clingen:#checks if the gene is on the clingen database
					if el[1] in clingen[value[0]][1]:#checks if the respective phenotype is on the clingen database
						ind=clingen[value[0]][1].index(el[1])
						cling='=HYPERLINK("'+clingen[value[0]][3][ind]+'","'+clingen[value[0]][2][ind]+'")'#prepares the text for the clingen part
				entries.append(['=HYPERLINK("http://omim.org/entry/'+el[0]+'","'+descr+'")','=HYPERLINK("http://omim.org/entry/'+el[0]+'","'+omim+'")', ddg,cling])#and prepares the link for omim with the text
				omim="ND"
				descr="ND"
				ddg="ND"
				cling="ND"
		if value[0] in dd2p and 'No disease mim' in dd2p[value[0]][1]:#if there is some phenotypes that appears only on DDG2P
			ind=[i for i, e in enumerate(dd2p[value[0]][1]) if e == "No disease mim"]#get the index number of these phenotypes
			for el in ind:#runs trough them
				ddg=dd2p[value[0]][0][el]#prepares the ddg2p part of the text
				descr=dd2p[value[0]][2][el]#prepares the phenotype description based on the DDG2P description, since OMIm is not available
				entries.append([dd2p[value[0]][-1][el].replace("SUBSTITUTE",descr), omim, dd2p[value[0]][-1][el].replace("SUBSTITUTE",ddg), cling])#makes the link, this ti or the HI/triplo fileme for the DDG2P database
				ddg="ND"
				descr="ND"
		if value[0] in clingen and 'No disease mim' in clingen[value[0]][1]:#if there is some phenotypes that appears only on clingen
			ind=[i for i, e in enumerate(clingen[value[0]][1]) if e == "No disease mim"]#get the index number of these phenotypes
			for el in ind:#runs trough them
				cling=clingen[value[0]][2][el]#prepares the clingen part of the text
				descr=clingen[value[0]][0][el]#prepares the phenotype description based on the clingen description, since OMIm is not available
				entries.append(['=HYPERLINK("'+clingen[value[0]][3][el]+'","'+descr+'")', omim, ddg, '=HYPERLINK("'+clingen[value[0]][3][el]+'","'+cling+'")'])#makes the link, this time for the clingen database
				cling="ND"
				descr="ND"
		if len(entries)==0:
			entries.append([descr, omim, ddg, cling])
		dicnew[key]=value
		dicnew[key][4]=entries#replaces the value[4] with this new list
		entries=[]
	return dicnew, forsub

def parse_first_search(response, filt, bp, chrr,key,numb, params):
	"""Parses the output from search, and returns two dictionaries
	one with the genomic elements disrupted by the breakpoint, as
	disrupted[gene/lincRNA..]=[chr:start-end, description/lincRNA_Ensembl_entry_name, strand, gene biotype, ensemblID]
	and another with all the genomic elements
	dic[gene]=[ch:start-end, description, strand, NOOMIM/OMIM ID, [phen/NOOMIM], gene biotype, ensembl ID]"""
	els=params["els"]
	dic={}
	disrupted=0
	aa={}
	for index, row in response.iterrows():
		print(row)
		if type(row[els[2]]) is str:
			s=int(row[els[0]])
			e=int(row[els[1]])
			if s<int(float(bp)) and e>int(float(bp)):
				disrupted+=1
			if row[els[2]].split(".")[0] not in aa:
				omid, pheno=marrvel_input.get_omim(row[els[2]].split(".")[0])
				if omid!="NOOMIM" or (row[els[3]]=="lincRNA" or row[els[3]]=="protein_coding") or (s<int(float(bp)) and e>int(float(bp))) or len(marrvel_input.read_gtex_loops(row[els[2]].split(".")[0], params["gtex"], "gtex",0))>0:
					dic[int(row[els[0]])]=[row[els[2]].split(".")[0],"chr"+chrr+":"+'{:0,}'.format(s)+"-"+'{:0,}'.format(e), row[els[5]], omid,pheno, row[els[3]], row[els[4]].strip(), [str(row[els[-1]]).strip()]]
				aa[row[els[2]].split(".")[0]]=int(row[els[0]])
			elif row[els[2]].split(".")[0] in aa :
				if aa[row[els[2]].split(".")[0]] in dic:
					if str(row[els[-1]]).strip() not in dic[aa[row[els[2]].split(".")[0]]][-1]:
						dic[aa[row[els[2]].split(".")[0]]][-1].append(str(row[els[-1]]).strip())
	dic, forsub=make_phen_nomenclature(dic, params)
	if int(filt["start"][0])<=int(float(bp)) and int(filt["end"][0])>=int(float(bp)):
		bb=int(float(bp))
		dic[int(float(bp))]=["Breakpoint"+numb, "Chr"+chrr+":"+'{:0,}'.format(bb), "", "","NOOMIM", "", ""]
	od = OrderedDict(sorted(dic.items()))
	final=OrderedDict()
	for key, value in od.items():
		final[value[0]]=value[1:]
	return final,disrupted, forsub

def interr(filt1,params):
	"""uses the search method to search for the disrupted genomic elements and returns a dictionary with
	the exons, as dic[ensemble_transcript_id]=[exonstart,exonend,exonstart,exonend...,transcript_name, gene_name, strand]"""
	dic={}
	fstep=search(filt1,params["at2"], params)
	els2=params["els2"]
	for index, row in fstep.iterrows():
		if row[els2[0]] not in dic and row[els2[1]]==row[els2[2]]:
			dic[row[els2[0]]]=[[row[els2[3]],row[els2[4]]], row[els2[5]], row[els2[-1]]-row[els2[-2]]]
		elif row[els2[0]] in dic and row[els2[1]]==row[els2[2]] and row[els2[-3]]!=dic[row[els2[0]]][1]:
			if int(row[els2[-1]]-row[els2[-2]])>int(dic[row[els2[0]]][-1]):
				dic[row[els2[0]]]=[[row[els2[3]],row[els2[4]]], row[els2[-3]], row[els2[-1]]-row[els2[-2]]]
		elif row[els2[0]] in dic and row[els2[1]]==row[els2[2]] and row[els2[-3]]==dic[row[els2[0]]][1]:
				dic[row[els2[0]]][0].append(row[els2[3]])
				dic[row[els2[0]]][0].append(row[els2[4]])
	new_dic={}
	for key,value in dic.items():
		value[0].sort()
		new_dic[value[-2]]=value[0]
	sstep=search({'link_ensembl_transcript_stable_id':list(new_dic.keys())}, params["at3"], params)
	els3=params["els3"]
	for index, row in sstep.iterrows():
		if row[els3[0]] in new_dic:
			new_dic[row[els3[0]]].append(row[els3[1]])
			new_dic[row[els3[0]]].append(row[els3[2]])
			new_dic[row[els3[0]]].append(row[els3[3]])
	return new_dic


def ivsreport(dic_pos, i0bp, i0bp2):
	"""Uses the dictionary from interr and the breakpoint to retrive a dictionary with the specific position of the
	breakpoint in the disrupted gene. Returns a dictionary as dic[gene name]=[Exon 1 - IVS7, IVS7 - Exon 9]"""
	bp=float(i0bp)
	bp2=""
	if i0bp2!="":
		bp2=float(i0bp2)
	dic_text={}
	sz=0
	temppy=0
	for key, value in dic_pos.items():
		if int(value[-4])>=int(bp) and int(value[0])<=int(bp):
			el=0
			sz=len(value)
			while el<len(value[:-4]):
				if int(bp)>=int(value[el]) and int(bp)<int(value[el+1]):
					if el%2==1:#uneven - intron
						aai="IVS "
					else:
						aai="Exon "
					if value[-1]=="1":
						st="; SS"
						numb=int(el/2+1)
					else:
						st="; AS"
						numb=int((len(value[:-3])-el)/2)
					if bp2!="":
						if (bp2<bp and value[-1]=="1") or (bp2>bp and value[-1]=="-1"):#se o bp2 aparecer primeiro que o bp1
							aa=["5'UTR-"+aai+str(numb)+st, bp]
						else:#o bp aparece primeiro que o bp2
							aa=[aai+str(numb)+"-3'UTR"+st, bp]							
					else:
						aa=[aai+str(numb)+st, bp]
					if type(value[-2]) is str:
						if value[-2].split(".")[0] not in dic_text or sz>temppy:
							temppy=sz
							dic_text[value[-2].split(".")[0]]=aa
				el+=1
	return dic_text

##################PREPARES THE EXEL FILE OF OUTPUT AND FORMATS THE CONTENT###################################

def prepare_exel(newname, first, namee, coments, params):
	"""makes the header of the new exel"""
	global ws2
	global wb
	if first==True and params["tt"]=="Spec_Rg":
		wb=Workbook()
		ws2=wb.active
		ws2.title =namee
	elif first==True:
		wb=Workbook()
		ws2=wb.active
		ws2.title =namee+"_rearrangement_A"
	else:
		ws2=wb.create_sheet()
		ws2.title=namee+"_rearrangement_B"#'=HYPERLINK("http://www.genecards.org/cgi-bin/carddisp.pl?gene='+dic[el][-1]+'", "'+el+papp+'")'
	dat= time.strftime("%d-%m-%Y")
	aa=['SVInterpreter', '', '','', '', dat]#https://gnomad.broadinstitute.org/
	ws2.append(aa)
	ws2.append([])
	ws2.append(newname)
	ws2.append([""])#linha a ser completada pela legenda que é feita aqui
	ws2.append(['Fields are in bold if: Genes are present in any panel from PanelAPP, HI%<10%, triplo. score = 3, o/e score ≤0.3, clustered interactions of GeneHancer are disrupted by the breakpoint, chromatin Loops are disrupted by the breakpoint; OMIM inheritance overlaping the inputed inheritance; SNPs with a p-value ≤5.0E-7.'])
	if params["tt"]=="Spec_Rg":
		ws2.append(['Region with yellow background is the one indicated for analysis.'])
	elif params["tt"]=="Duplication":
		ws2.append(['Regions with blue background are duplicated.'])
	else:
		ws2.append(['Regions with red background are deleted.'])
	ws2.append([])
	ws2.append(['Genes and intergenic regions', '', '', '','','','','','','','', 'Clinical Phenotype','','','','', 'Gene fusion in cancer','','Gene-phenotype/disease associations and animal models','','','','', 'Infertility','GWAS data', 'Bibliography', 'Only for CNVs'])
	ws2.append(['=HYPERLINK("https://www.genecards.org/","GeneCard")', '=HYPERLINK("https://www.coriell.org/1/NIGMS/Collections/ACMG-59-Genes","ACMG")', '=HYPERLINK("https://www.ensembl.org/index.html","Ensembl")', '=HYPERLINK("https://omim.org/","OMIM")', '=HYPERLINK("https://dosage.clinicalgenome.org/","Dosage Map")','=HYPERLINK("https://gnomad.broadinstitute.org/","GnomAD")', 
	'=HYPERLINK("https://www.uniprot.org/","Uniprot")', '=HYPERLINK("https://omim.org/","OMIM")','=HYPERLINK("https://www.gtexportal.org/home/index.html","Gtext expression")', '=HYPERLINK("https://genome.ucsc.edu/cgi-bin/hgTrackUi?db=hg19&g=geneHancer","GeneHancer")', '=HYPERLINK("http://3dgenome.fsm.northwestern.edu/publications.html","Loops")', 'Phenotype', '=HYPERLINK("https://omim.org/","OMIM")', 
	'=HYPERLINK("https://www.ebi.ac.uk/gene2phenotype/disclaimer","DDG2P")', '=HYPERLINK("https://clinicalgenome.org/","ClinGen")', '=HYPERLINK("https://www.ncbi.nlm.nih.gov/pmc/articles/PMC4321842/","Phenotype similarity")','=HYPERLINK("http://atlasgeneticsoncology.org/","AGCOH")','=HYPERLINK("http://atlasgeneticsoncology.org/","Oncology Atlas")', '=HYPERLINK("https://wormbase.org//#012-34-5","Wormbase")', 
	'=HYPERLINK("https://flybase.org","FlyBase")', '=HYPERLINK("http://www.informatics.jax.org/","MGI")', '=HYPERLINK("https://rgd.mcw.edu/rgdweb/homepage/","RGD")', '=HYPERLINK("https://zfin.org/","Zfin")', '=HYPERLINK("https://pubmed.ncbi.nlm.nih.gov/30865283/","Infertility study")', '=HYPERLINK("https://www.ebi.ac.uk/gwas/","GWAS Catalog")','=HYPERLINK("https://pubmed.ncbi.nlm.nih.gov/","References")', '=HYPERLINK("http://dgv.tcag.ca/dgv/app/home","CNV database overlap")'])
	ws2.append(['Genomic elements; Pannels from PannelApp','Actionable Genes (MAGs)', 'Breakpoint location;Genome strand','Gene ID', 'HI% ; Triplo', 'pLi; o/e', 'Protein entries', 'Function', 'Top 3 highest TPM (Total TPM; Mean TPM)', 'Clustered interactions','Chromatin Loops ','Assoc. Disorder',
	 'OMIM ID_Inh', 'DDG2P class.', 'clinGen class.', 'PhenSSc (P; MaxSSc)', 'Gene1 / Gene2 Cytoban', 'Organ: type, nr. cases', 'C.elegans', 'Drosophila', 'Mouse', 'Rat', 'Zebrafish', 'Disorder','SNPs - Genetic traits', 'PubMed Link', 'Best Hits'])
	#forhtml.write_start()
	szz={"A":[200,200], "B":[200,200], "C":[100,200], "D":[100,200], "E":[100,200], "F":[200,200], "G":[100,100], "H":[200,200], "I":[200,200], "J":[300,200],"K":[200,200], "L":[150,100], "M":[100,200], "N":[200,200], "O":[200,200], "P":[200,300], "Q":[200,200], "R":[200,200], "S":[200,200], "T":[200,200], "U":[200,200], "V":[200,200], "W":[200,200], "X":[200,200], "Y":[200,200],"Z":[200,200],"AA":[200,200]}
	for key,value in coments.items():
		coment=Comment(value, "Info:")
		coment.width=szz[key][0]
		coment.height=szz[key][1]
		ws2[key+"10"].comment=coment

def check_bands(text, name):
	d=text.split("(")
	chrA=d[-2].split(";")[0]
	ps=[i for i, ltr in enumerate(d[-1]) if ltr == "p"]
	qs=[i for i, ltr in enumerate(d[-1]) if ltr == "q"]
	if len(ps)>0 and len(qs)>0:
		if qs[0]>ps[0]:
			if chrA+"p" in name:
				return ["p","q", "first"]
			else:
				return ["p","q", "second"]
		else:
			if chrA+"q" in name:
				return ["q","p", "first"]
			else:
				return ["q","p", "second"]
	elif len(ps)>0:
		if name.startswith(chrA+"p"):
			return ["p","p", "first"]
		else:
			return ["p","p", "second"]
	else:
		if name.startswith(chrA+"q"):
			return ["q","q", "first"]
		else:
			return ["q","q", "second"]

def make_background(ws2, params):
	"""see which cells need backgroud and apply it"""
	tt=check_points(ws2)
	bands=[]
	if params["tt"]=="Balanced_translocation":
		bands=check_bands(ws2["A3"].value, ws2.title)
	print(bands)
	i=11
	aa=False
	ist=False
	forder=False
	lets=["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L","M","N","O","P","Q","R","S","T","U","V","W","X", "Y", "Z","AA"]
	while i<=ws2.max_row:
		bb=ws2["A"+str(i)].value
		if bb!=None:
			if bb.startswith("pter"):
				ist=True
			elif (bb.startswith("Breakpoint") or bb.startswith("chr") or bb.startswith("g.") or "Query Region" in bb) and aa==False and tt!=1:
				aa=True
				forder=True
				if ist==True:
					ws2.delete_rows(i,1)
					i-=1
			elif (bb.startswith("Breakpoint") or bb.startswith("chr") or bb.startswith("g.") or "Query Region" in bb) and aa==False and tt==1:
				forder=True
			elif (bb.startswith("Breakpoint")==False and bb.startswith("chr")==False and bb.startswith("g.")==False and "Query Region" not in bb) and aa==True:
				for el in lets:
					ff=ws2[el+str(i)]
					if params["tt"]=="Duplication":
						ff.fill=PatternFill("solid", fgColor="CCCEFF")#azul
					elif params["tt"]=="Spec_Rg":
						ff.fill=PatternFill("solid", fgColor="FCFFBA")#amarelo
					else:
						ff.fill=PatternFill("solid", fgColor="FFCCCC")#vermelho
			elif (bb.startswith("Breakpoint") or bb.startswith("chr") or bb.startswith("g.") or "Query Region" in bb) and aa==True and tt!=1:
				if ist==False:
					ws2.delete_rows(i,1)
					i-=1
					ist=False
				aa=False
			elif len(bands)>0:
				if (bands[2]=="first" and bands[0]=="p" and forder==False) or (bands[2]=="first" and bands[0]=="q" and forder==True) or (bands[2]=="second" and bands[1]=="p" and forder==True) or (bands[2]=="second" and bands[1]=="q" and forder==False):
					if bb.startswith("Analysis performed wit")==False:
						for el in lets:
							ff=ws2[el+str(i)]
							ff.fill=PatternFill("solid", fgColor="FCFFBA")#amarelo trans
				else:
					if bb.startswith("Analysis performed wit")==False:
						for el in lets:
							ff=ws2[el+str(i)]
							ff.fill=PatternFill("solid", fgColor="F4F5E1")#cinza trans
		i+=1

def make_format_ws2(ws2, flag, is_region, params):
	"""changes fonts, read sizes, bold, italic, colors and cell limits of the table
	strutures almost all the graphic vizualization of the output"""
	"""formats the exel file to a table format"""
	if (flag==True and is_region==False) or params["tt"]=="Balanced_translocation":
		make_background(ws2, params)
	thick_border=Border(bottom=Side(style='thick'))
	top_border=Border(top=Side(style='thin'),bottom=Side(style='thin'))
	maxborder=Border(top=Side(style='thin'))###
	col_border=Border(left=Side(style='thin'))
	bottom_thin=Border(bottom=Side(style="thin"))
	a1=ws2["A1"]
	a1.font=Font(bold=True, size=12, color="808080")
	a1=ws2["D1"]
	a1.font=Font(bold=True, size=12, color="808080")
	a1=ws2["F1"]
	a1.font=Font(bold=True, size=11, color="0000CC")
	a1=ws2['A3']
	a1.font=Font(size=11)
	a1=ws2['A4']
	a1.font=Font(name="Arial Narrow", size=10)
	a1=ws2['A5']
	a1.font=Font(name="Arial Narrow", size=10)
	a1=ws2['A6']
	a1.font=Font(name="Arial Narrow", size=10)
	ws2.merge_cells("A3:H3")
	ws2.merge_cells("A4:Y4")
	ws2.merge_cells("A5:Y5")
	i=11
	yy=0
	lets=["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L","M","N","O","P","Q","R","S","T","U","V","W","X", "Y", "Z", "AA"]
	with_vorder=["L", "Q", "S", "X","Y","Z","AA"]
	with_bold=["E", "F", "J", "K","M", "Y"]
	buish=["D","G", "L", "N", "O", "Q", "R", "S", "T", "U", "V", "W", "Z", "AA"]
	widths={"A":12.29, "B":8.43, "C":8.43, "D":6.57, "E":9.00, "F":12.57, "G":12.57,"H":25.86, "I":24.00, "J":21.14, "K":65.43, "L":18.14, "M":11.43, "N":11.71, "O":11.57, "P":16.57, "Q":24.71, "R":25.00, "S":16.43,"T":16.43,"U":16.43,"V":16.43,"W":16.43,"X":16.43, "Y":59.14,"Z":10.57,"AA":20.29}
	aftertad=False
	need_another=False
	##
	ws2.merge_cells("A8:K8")
	ws2.merge_cells("L8:P8")
	ws2.merge_cells("Q8:R8")
	ws2.merge_cells("S8:W8")
	n=ws2["A8"]
	n.font=Font(size=12, bold=True, name="Arial Narrow")
	n=ws2["L8"]
	n.font=Font(size=12, bold=True, name="Arial Narrow")
	n=ws2["Q8"]
	n.font=Font(size=12, bold=True, name="Arial Narrow")
	n=ws2["S8"]
	n.font=Font(size=12, bold=True, name="Arial Narrow")
	n=ws2["X8"]
	n.font=Font(size=12, bold=True, name="Arial Narrow")
	n=ws2["Y8"]
	n.font=Font(size=12, bold=True, name="Arial Narrow")
	n=ws2["Z8"]
	n.font=Font(size=12, bold=True, name="Arial Narrow")
	n=ws2["AA8"]
	n.font=Font(size=12, bold=True, name="Arial Narrow")
	ws2.row_dimensions[10].height = 50.25
	for el in lets:
		if el  not in with_vorder:
			n=ws2[el+"10"]
			n.border=Border(bottom=Side(style='thick'))
			n.alignment=Alignment(wrap_text=True)
			n=ws2[el+"8"]
			n.border=Border(bottom=Side(style='thin'))
		else:
			n=ws2[el+"10"]
			n.border=Border(bottom=Side(style='thick'),left=Side(style='thin'))
			n.alignment=Alignment(wrap_text=True)
			n=ws2[el+"8"]
			n.border=Border(left=Side(style='thin'),bottom=Side(style='thin'))
		n=ws2[el+"7"]
		n.border=Border(bottom=Side(style='thick'))
		n.font=Font(size=11, bold=True, name="Arial Narrow")
		n=ws2[el+"9"]
		if "HYPERLINK" in n.value:
			n.font=Font(size=11, bold=True, name="Arial Narrow", color="0000CC")
		else:
			n.font=Font(size=11, bold=True, name="Arial Narrow")
		n=ws2[el+"10"]
		n.font=Font(size=11, bold=True, name="Arial Narrow")
	for key,value in widths.items():
		ws2.column_dimensions[key].width = value
	while i<=ws2.max_row:
		for l in lets:
			gg=ws2[l+str(i)]
			gg.alignment=Alignment(horizontal="left", vertical="top", wrap_text=True)
			if l not in with_bold:
				gg.font=Font(name="Arial Narrow", size=10)
		n=ws2["E"+str(i)]
		index="A"+str(i)
		bb=ws2[index].value
		if bb!=None:
			if "Telomere" in bb or "Centromere" in bb:
				ws2[index].font=Font(bold=True)
				for tt in lets:
					n=ws2[tt+str(i)]
					n.border=Border(top=Side(style="thick"))
			elif ("Intergenic" in bb and "regions" not in bb) or "None" in bb:
				ws2.merge_cells("A"+str(i)+":C"+str(i))
				for tt in lets:
					n=ws2[tt+str(i)]
					n.font=Font(name="Arial Narrow", size=10)
					if aftertad==False:
						if tt in with_vorder:
							n.border=Border(top=Side(style='thin'),left=Side(style='thin'))
						else:
							#n.border=Border(top=Side(style='thin'),left=Side(style='thin'))
							n.border=maxborder
				aftertad=False
			elif ("TAD" in bb and "genome" not in bb) or ("Deleted" in bb and "genome" not in bb) or bb.startswith("qter") or bb.startswith("pter"):
				ws2.merge_cells("A"+str(i)+":G"+str(i))
				n=ws2["A"+str(i)]
				n.font=Font(bold=True, size=12, name="Arial Narrow")
				aftertad=True
				if yy==0:
					for el in lets:
						n=ws2[el+str(i)]
						n.border=Border(top=Side(style='thick'), bottom=Side(style='dashed'))
						yy=1
				else:
					for el in lets:
						n=ws2[el+str(i)]
						n.border=Border(top=Side(style='thick'), bottom=Side(style='dashed'))					
			elif bb.startswith("Breakpoint") or bb.startswith("g.") or "del" in bb or "dup" in bb or bb.startswith("Query Region"):
				ws2.merge_cells("A"+str(i)+":G"+str(i))		
				n=ws2["A"+str(i)]
				if params["tt"]=="Duplication":
					n.font=Font(bold=True, size=11, name="Arial Narrow", color="4249FC")#breakpoint a vermel
				elif params["tt"]=="Deletion":
					n.font=Font(bold=True, size=11, name="Arial Narrow", color="FC3232")#breakpoint a vermel
				elif params["tt"]=="Spec_Rg":
					n.font=Font(bold=True, size=11, name="Arial Narrow", color="F7D943")#breakpoint a vermel
				elif params["tt"]=="Inversion":
					n.font=Font(bold=True, size=11, name="Arial Narrow", color="FC42E7")#breakpoint a vermel
				else:
					n.font=Font(bold=True, size=11, name="Arial Narrow", color="009900")#breakpoint a vermelho
				if need_another==False and aftertad==False:
					for tt in lets:
						n=ws2[tt+str(i)]
						if tt not in with_vorder:
							n.border=maxborder
						else:
							n.border=Border(top=Side(style='thin'),left=Side(style='thin'))
				else:
					for tt in lets:
						n=ws2[tt+str(i)]
						n.border=Border(bottom=Side(style='thin'))
					aftertad=False
					need_another=False
			elif "genome" in bb:
				ws2.merge_cells("A"+str(i)+":Y"+str(i))
				for el in lets:
					n=ws2[el+str(i)]
					n.border=Border(top=Side(style="thick"))
				n=ws2["A"+str(i)]
				n.font=Font(size=11, name="Arial Narrow")
				break
			elif bb!="" and ws2["G"+str(i)].value!=None:
				if aftertad==False:
					for tyt in lets:
						njj=ws2[tyt+str(i)]
						if tyt not in with_vorder and ws2["D"+str(i)].value!="":
							njj.border=maxborder
						if tyt not in with_bold and ws2["D"+str(i)].value!="":
							if tyt=="C" and ws2["C"+str(i)].value!=None:
								if tyt=="C" and ("IVS" in  ws2["C"+str(i)].value or "Exon" in ws2["C"+str(i)].value):
									ws2["C"+str(i)].font=Font(size=10, name="Arial Narrow", bold=True, color="009900")
								else:
									njj.font=Font(size=10, name="Arial Narrow")
							else:
								njj.font=Font(size=10, name="Arial Narrow")
				n=ws2["A"+str(i)]
				ghg=n.value
				if len(ghg.split('"'))>2:
					if "*" not in ghg:
						n.font=Font(color="0000CC", italic=True, size=10, name="Arial Narrow")
					else:
						n.font=Font(color="0000CC", italic=True, size=10, bold=True, name="Arial Narrow")
				else:
					n.font=Font(name="Arial Narrow", size=10)
				aftertad=False
			for ttt in buish:
				if ws2[ttt+str(i)].value !=None:
					if "HYPERLINK" in ws2[ttt+str(i)].value:
						ws2[ttt+str(i)].font=Font(color="0000CC", size=10, name="Arial Narrow")
					else:
						ll=ws2["A"+str(i)].value
						if "TAD" not in ll and "Deleted" not in ll  and ll.startswith("qter")==False and ll.startswith("pter")==False and ll.startswith("der")==False and ll.startswith("g.")==False and "del" not in ll and "dup" not in ll:
							ws2[ttt+str(i)].font=Font(size=10, name="Arial Narrow")
		i+=1
	final_touch(ws2, i, with_vorder, col_border, with_bold, params)
	ws2.freeze_panes = 'B11'  # this will freeze the first column
	

def make_ACMG_table_format(ws2, i):
	i+=2
	cols=["A", "B", "C"]
	while i<=ws2.max_row:
		index="A"+str(i)
		bb=ws2[index].value
		if bb!=None:
			if bb.startswith("Partial"):
				ws2[index].font=Font(name="Arial Narrow", size=12, bold=True)
			elif "HYPERLINK" in bb:
				ws2[index].font=Font(color="0000CC", size=10, name="Arial Narrow")
			elif bb.startswith("Parameter"):
				for el in cols:
					ws2[el+str(i)].font=Font(name="Arial Narrow", size=12, bold=True)
					ws2[el+str(i)].border=Border(top=Side(style='thick'),bottom=Side(style='thick'))
			elif bb.startswith("This"):
				ws2[index].font=Font(size=10, name="Arial Narrow")
			else:
				for el in cols:
						ws2[el+str(i)].font=Font(name="Arial Narrow", size=10)
						if i==ws2.max_row:
							ws2[el+str(i)].border=Border(bottom=Side(style='thick'))
		i+=1

def final_touch(ws2, i, with_vorder, col_border, with_bold, params):
	for countt in range(11,i):
		for colm in with_vorder:
			test2=ws2["A"+str(countt)].value
			if test2!=None:
				if "g." not in test2:
					if "Intergenic" in test2:
						ws2[colm+str(countt)].border=Border(top=Side(style='thin'),left=Side(style='thin'))
				if test2!=None and ws2["D"+str(countt)].value!=None:
					if (len(test2)>1 and len(ws2["D"+str(countt)].value)>1) or "Intergenic" in test2:
						ws2[colm+str(countt)].border=Border(top=Side(style='thin'),left=Side(style='thin'))
					else:
						ws2[colm+str(countt)].border=col_border
			else:
				if "Intergenic" in test2:
					ws2[colm+str(countt)].border=Border(top=Side(style='thin'),left=Side(style='thin'))
				else:
					ws2[colm+str(countt)].border=col_border
		for col in with_bold:
			test=ws2[col+str(countt)].value
			if test!=None:
				if (col!="L" and test.startswith("b ")) or ("b SNP" in test and col=="Y"):
					temporal2=test.replace("b ", "", 1)
					ws2[col+str(countt)].font=Font(name="Arial Narrow", size=10, bold=True, color="009900")
					ws2[col+str(countt)].value=temporal2
				elif "b "+params["input_inh"] in test and col=="L":
					temporal2=test.replace("b ", "", 1)
					ws2[col+str(countt)].font=Font(name="Arial Narrow", size=10, bold=True, color="0000CC")
					ws2[col+str(countt)].value=temporal2
				else:
					if col!="L" or test=="ND":
						ws2[col+str(countt)].font=Font(name="Arial Narrow", size=10)
					else:
						ws2[col+str(countt)].font=Font(name="Arial Narrow", size=10, color="0000CC")

########################Executes the steps (read the input, organize the tasks, write output)#############################

def final_things_sub(ch, value, bp, key, numb, bp2, params):
	"""acts as part of the final_things function. This function executes the first search using the breakpoint,
	cromossome and TADs. Returns the dictionary with the genes and respective information and a dictionary with the
	genes/genomic elements disrupted by the breakpoint."""
	interromped={}
	filt1= {'chromosome_name': [ch], 'start': [str(value[0])], 'end':[str(value[1])]}#alt
	dic,disrupted, forsub=parse_first_search(search(filt1, params["at1"], params),filt1,bp,ch,key, numb, params)#alt
	dic_text={}
	if disrupted!=0:
		dic_text=ivsreport(interr({'chromosome_name': [ch], 'start': [bp], 'end':[str(int(float(bp))+1)]},params), bp, bp2)
	return dic, dic_text, forsub

def get_info_forsub(forsub, papp):
	temporal=papp.replace("*","")
	temporal2=temporal.replace(";",",")
	for eleap in temporal.split(","):
		forsub.add(eleap)
	gggg=""
	if "***" in papp:
		forsub.add("***")
	elif "**" in papp:
		forsub.add("**")
	elif "*" in papp:
		forsub.add("*")
	return forsub

def confirm_disrupted(el, bps):
	if type(bps) is str:
		bps=[bps]
	if el!="nd" and el!="":
		for bp in bps:
			if "&&" not in el:
				l=el.split("-")
				g=l[0].split(":")
				if int(g[1].replace(",",""))<int(float(bp)) and int(l[1].replace(",",""))>int(float(bp)):
					el="b "+el+" (disrupted)"
					break
			else:
				l=el.split("&&")
				start=l[0].split(":")[1][:-2]
				s=start.split("-")
				end=l[1].split(":")[1][:-2]
				e=end.split("-")
				if (int(s[0].replace(",",""))<int(float(bp)) and int(e[1].replace(",",""))>int(float(bp))) or (int(e[0].replace(",",""))>int(float(bp)) and int(s[1].replace(",",""))<int(float(bp))):
					el="b "+el+" (disrupted)"
					break
	return el

def is_inside(elie, dic, pp, br):
	if pp!="Deletion" and pp!="Duplication":
		return False
	else:
		de=dic[elie][0].split(":")[1]
		dd=de.split("-")[0]
		if int(dd.replace(",",""))>=int(br[0]) and int(dd.replace(",",""))<int(br[1]):
			return True
		else:
			return False

def parse_inf(oe, ell, synn):
	"""reads the oe file or the HI/triplo file or the genehancer file
	, and returns the oe value for the
	gene indicated by the variable ell."""
	f=open(oe, encoding="UTF-8")
	f.readline()
	nn="nd"
	if type(synn) is list:
		synn.append(ell)
	else:
		synn=[ell]
	for el in f:
		line=el.split("\t")
		if line[0] in synn:
				nn=line[-1].strip()#returns the oe with the confidence interval, in all cases
				break
	f.close()
	return nn



def make_report_table(el,dic, dic_text, chrrr, name, tad, params, forsub, bp, br, band):
	"""called to organize the output, and write it. 
	This function interacts directly with the output file"""
	if el!="None":
		infert=parse_inf(params["infertility"], el, dic[el][-1])
	else:
		infert="nd"
	if el!="None" and el.startswith("chr")==False and el.startswith("Chr")==False and "Breakpoint" not in el not in dic_text and is_inside(el, dic, params["tt"], br)==False:
		tonco=panelapp.tiny_transform(params["tonco"])
	else:
		tonco=panelapp.read_and_transform(params["onco"])
	acmg=parse_acmg(params["acmg"])
	if el in acmg:
		tto=el
	else:
		tto="ND"
	if el not in tonco:
		tonco[el]=[["ND","ND"]]
	dicd=list(dic_text.keys())
	if el=="None":
		ws2.append(["None"])
	elif el.startswith("Chr") or el.startswith("chr"):
		snps=read_snps_new.select_all_snps(params["snps"][-1], dic[el][0], dic[el][1], dic[el][2],params["snps"][1])
		if len(snps)>0:
			if len(snps)==1:
				snps_numb=list(snps.keys())
			else:
				snps_numb=list(snps.keys())
				snps_numb.sort(reverse=True)
			fff=0
			while fff<len(snps_numb):
				snf=""
				if fff<len(snps_numb):
					snf="# "+str(snps_numb[fff])+" SNPs - "+snps[snps_numb[fff]]
				if fff==0:
					ws2.append(["Intergenic - "+el, "","","", "","","","","", "","","","","", "","","","","", "","","","","",snf, "", ""])
					#forhtml.write_html("line", ["Intergenic - "+el, " "," ", " "," "," "," "," ", " "," "," "," "," ", " "," "," "," "," ", " "," "," "," ",snf, " ", " "], "tool_temp_data/btemp.html")
				else:
					ws2.append(["","", "","","","","","", "","","","","", "","","","","", "","","","", "","",snf, ""])
					#forhtml.write_html("line", [" "," ", " "," "," "," "," ", " "," "," "," "," ", " "," "," "," "," ", " "," "," "," ", ' ',snf, " "], "tool_temp_data/btemp.html")
				fff+=1
	elif el.startswith("Breakpoint"):
		cnv_hits=params["hits"]
		if len(cnv_hits)>0 and params["is_hit_in"]==False:
			aa=0
			while aa<len(cnv_hits):
				if aa==0:
					ws2.append([name,"","","","","","","","","","","","","","","","","","","","","","","","",marrvel_input.make_biblio_link_DD(params["tt"],band),cnv_hits[aa]])
					#forhtml.write_html("line", ["b "+name," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," ",cnv_hits[aa]], "tool_temp_data/btemp.html")####
				else:
					ws2.append(["","","","","","","","","","","","","","","","","","","","","","","","","","",cnv_hits[aa]])
					#forhtml.write_html("line", [" "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," ",cnv_hits[aa]], "tool_temp_data/btemp.html")####
				aa+=1
			params["is_hit_in"]=True
		else:
			ws2.append([name, "","","","","","","","","","","","","","","","","","","","","","","","",marrvel_input.make_biblio_link_DD(params["tt"],band),""])
			#forhtml.write_html("line", ["b "+name," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "," "], "tool_temp_data/btemp.html")####
	else:
		if el in dic_text:
			strand_location=dic_text[el][0]
		else:
			if dic[el][1]=="-1":
				strand_location="AS"
			else:
				strand_location="SS"
		papp=parse_oe(params["panel"], el)
		if papp!="nd":
			forsub=get_info_forsub(forsub, papp)
		else:
			papp=""
		genes_snps, nsnps=read_snps_new.read_gene_traits(params["snps"][0], dic[el][-2], params["snps"][1], tad[0], tad[1])
		if len(el.split(" "))>1:
			gg=el.split(" ")[1]
			el=gg
		genecard='=HYPERLINK("http://www.genecards.org/cgi-bin/carddisp.pl?gene='+dic[el][-2]+'", "'+el+papp+'")'
		if len(genecard)!=0:
			pli=parse_oe(params["oe"], el)#oe
			hi=parse_oe(params["hi"], el)#hi/triplo info
			params["oes"][el]=[hi,pli]
			genehancer=confirm_disrupted(parse_oe(params["genehancer"], dic[el][-2]),bp)#genehancer data
			if genehancer=="nd":
				genehancer=confirm_disrupted(parse_oe(params["genehancer"], el),bp)#genehancer data
			animals=marrvel_input.get_all_animal_models(el,params,dic[el][-1])#animal models on a list
			aa=True
			fgg=0
			top3gtex=marrvel_input.read_gtex_loops(dic[el][-2], params["gtex"], "gtex", chrrr)
			loops=marrvel_input.read_gtex_loops(el, params["tad"][-1], "loops", chrrr)
			if len(dic[el][3])!=0 or len(nsnps)!=0 or len(tonco[el])!=0 or len(top3gtex)!=0 or len(loops)!=0 or len(animals[0])!=0 or len(animals[1])!=0 or len(animals[2])!=0 or len(animals[3])!=0 or len(animals[4])!=0:
				counteer=max([len(dic[el][3]), len(nsnps), len(tonco[el]), len(top3gtex), len(loops), len(animals[0]),len(animals[1]),len(animals[2]),len(animals[3]),len(animals[4])])
				while fgg<counteer:
					if aa==True:
						lin=["ND", "ND", "ND", "ND"]
						oonco=["ND","ND"]
						lop="ND"
						ggtex="ND"
						worm="ND"
						dros="ND"
						mouse="ND"
						rat="ND"
						zfish="ND"
					if aa==False:
						lin=[" "," "," "," "]
						oonco=[" "," "]
						ggtex=" "
						lop=" "
						worm=""
						dros=""
						mouse=""
						rat=""
						zfish=""
					snpss=" "
					if fgg<len(dic[el][3]):
						lin=dic[el][3][fgg]
					if fgg<len(nsnps):
						snpss="# "+str(nsnps[fgg])+" SNPs - "+genes_snps[str(nsnps[fgg])]
					if fgg<len(tonco[el]):
						oonco=tonco[el][fgg]
					if fgg<len(top3gtex):
						ggtex=top3gtex[fgg]
					if fgg<len(loops):
						lop=confirm_disrupted(loops[fgg], bp)
					if fgg<len(animals[0]):
						worm=animals[0][fgg]
					if fgg<len(animals[1]):
						dros=animals[1][fgg]
					if fgg<len(animals[2]):
						mouse=animals[2][fgg]
					if fgg<len(animals[3]):
						rat=animals[3][fgg]
					if fgg<len(animals[4]):
						zfish=animals[4][fgg]
					if aa==True:
						if dic[el][2]=="NOOMIM":
							ws2.append([genecard,tto,strand_location,"na",hi, pli, '=HYPERLINK("https://www.uniprot.org/uniprot/?query='+el+'&fil=organism%3A%22Homo+sapiens+%28Human%29+%5B9606%5D%22&sort=score","'+el+' Human Uniprot entry")', marrvel_input.get_description(el),ggtex,genehancer, lop, lin[0], lin[1], lin[2], lin[3],"ND", oonco[0],oonco[1], worm, dros, mouse, rat, zfish,infert, rep(snpss), marrvel_input.make_biblio_link(el, dic[el][-1])," "])
							#forhtml.write_html("line", [genecard,tto,strand_location,"na",hi, pli,marrvel_input.get_description(el),rep(ggtex),genehancer, rep(lop), rep(lin[0]), rep(lin[1]), rep(lin[2]), rep(lin[3]),"ND", rep(oonco[0]), rep(oonco[1]),  worm, dros, mouse, rat, zfish, snpss, marrvel_input.make_biblio_link(el)," "], "tool_temp_data/btemp.html")####
						else:
							if "_" in lin[1] and len(params["hpo_des"])>1:
								hpo_v=execute_hposim(params["hpo_des"],lin[1].split("_")[0])
							else:
								hpo_v="ND"
							ws2.append([genecard,tto,strand_location, '=HYPERLINK("https://omim.org/entry/'+dic[el][2]+'","'+dic[el][2]+'")',hi,pli,'=HYPERLINK("https://www.uniprot.org/uniprot/?query='+el+'&fil=organism%3A%22Homo+sapiens+%28Human%29+%5B9606%5D%22&sort=score","'+el+' Human Uniprot entry")',marrvel_input.get_description(el),rep(ggtex),genehancer,rep(lop), rep(lin[0]), rep(lin[1]), rep(lin[2]), rep(lin[3]), hpo_v, rep(oonco[0]), rep(oonco[1]), worm, dros, mouse, rat, zfish, infert, rep(snpss), marrvel_input.make_biblio_link(el, dic[el][-1])," "])
							#forhtml.write_html("line", [genecard,tto,strand_location, '=HYPERLINK("https://omim.org/entry/'+dic[el][2]+'","'+dic[el][2]+'")',hi,pli,marrvel_input.get_description(el),ggtex,genehancer,lop, lin[0], lin[1], lin[2], lin[3], hpo_v, oonco[0], oonco[1], worm, dros, mouse, rat, zfish, snpss, marrvel_input.make_biblio_link(el)," "], "tool_temp_data/btemp.html")####
						aa=False
						hjh=True
					else:
						if "_" in lin[1] and len(params["hpo_des"])>1:
							hpo_v=execute_hposim(params["hpo_des"],lin[1].split("_")[0])
						else:
							hpo_v="ND"
						if hjh==True and dic[el][-1]!=["nan"]:
							ws2.append(["("+",".join(dic[el][-1])+")","","","","","","", "", rep(ggtex), "", rep(lop), rep(lin[0]), rep(lin[1]), rep(lin[2]), rep(lin[3]), hpo_v, rep(oonco[0]), rep(oonco[1]),  worm, dros, mouse, rat, zfish, "",rep(snpss), "", ""])#tem de ser alterado para o numero
							hjh=False
						else:
							hjh=False
							ws2.append(["","","","","","","", "", rep(ggtex), "", rep(lop), rep(lin[0]), rep(lin[1]), rep(lin[2]), rep(lin[3]), hpo_v, rep(oonco[0]), rep(oonco[1]),  worm, dros, mouse, rat, zfish, "",rep(snpss), "", ""])#tem de ser alterado para o numero
						#forhtml.write_html("line", [" "," "," "," "," "," ", " ", ggtex, " ", lop, lin[0], lin[1], lin[2], lin[3], hpo_v,oonco[0], oonco[1],  worm, dros, mouse, rat, zfish, snpss, " ", " "] , "tool_temp_data/btemp.html")####
						#ws2.append([genecard,"strand_function", '=HYPERLINK("https://omim.org/entry/'+dic[el][2]+'","'+dic[el][2]+'")',hi,pli,"function","gtex","genehancer","loops", lin,"OMIM ID_Inh","ddg2p","clingen","hpo", oonco,"onco2","wormbase","flybase","mgi","rat","zebrafish", snpss, "pubmed","cnvs"])
					fgg+=1
			if hjh==True:
				ws2.append(["("+",".join(dic[el][-1])+")","","","","","","", "", "", "", "", "", "", "", "", "", "", "",  "","", "", "", "", "", "", "", ""])#tem de ser alterado para o numero

	make_sub(forsub)

def rep(v):
	if v==" ":
		return ""
	else:
		return v

def make_sub(forsub):
	inttext1=":"
	inttext2=":"
	text1="PanelAPP (Martin et al. 2019) data is showed next to the Gene name"
	text2=" Haploinsuficiency index (%HI) according to Decipher (Wright et al. 2016); Triplosensitivity score (Triplo) according to ClinGen: 0 - No evidence for dosage pathogenicity, 1 - Little evidence for dosage pathogenicity, 2 -	Some evidence for dosage pathogenicity, 3 - 	Sufficient evidence for dosage pathogenicity, 40 - Evidence suggests the gene is not dosage sensitive; observed vs. expected (oe) number of LoF variants and respective confidence interval according to gnomad (Karczewski et al. 2020); Tissue with the highest expression according to Gtex (Lonsdale et al. 2013), associated with the respective Transcripts Per Kilobase Million (TPM) and the sum of all TPM of all tissues analyzed; Clustered interactions of GeneHancer (Fishilevich et al. 2017) regulatory elements and genes; Chromatin loops identified on the reference cell line; OMIM phenotype ID, and respective inheritance"
	text3="Developmental Disorders Genotype-to-Phenotype database (DDG2P) classification of the phenotype(Wright et al. 2015); ClinGen classification of the phenotype (Heidi et al. 2015); Phenotype similarity analysis using the Human phenotype ontology (HPO) (ref. para o software de similaridade); Fusion gene found in cancer-derived rearrangement, the associated phenotype and the number of cases found (Huret et al. 2013, Mitleman 2020); Phenotypic characteristics identified on studies with animal models (wormbase 2020, Thurmond et al. 2019, Bult et al. 2019, Steen et al. 1999, Ruzicka et al. 2019); Genome wide association studies (GWAS) SNPs, associated pathologies and p-value (Buniello et al 2019)."
	ph=OrderedDict([("***","high evidence"), ("**", "moderate evidence"), ("*", "low evidence"),("A","Actionable information"), ("B","Cancer Programme"), ("C","Cardiovascular disorders"), ("D","Ciliopathies"), ("E","Dermatological disorders"),
	("F","Dysmorphic and congenital abnormality syndromes"), ( "G","Endocrine disorders"), ("H","Gastroenterological disorders"), ("I","Growth disorders"), ("J","Haematological and immunological disorders"),
	("K","Haematological disorders"), ("L","Hearing and ear disorders"), ("M","Metabolic disorders"), ("N","Neurology and neurodevelopmental disorders"), ("O","Ophthalmological disorders"),
	("P","Renal and urinary tract disorders"), ("Q","Respiratory disorders"), ("R", "Rheumatological disorders"),("S","Skeletal disorders"), ("T", "Tumour syndromes"), ("U", "Others")])
	inh=OrderedDict([("AD", "Autossomal Dominant"), ("AR", "Autossomal Recessive"), ("IC", "Isolated Cases"), ("SMu", "Somatic Mutation"), ("Mu", "Multifactorial"), ("Mi", "Mitochondrial"), ("SMo", "Somatic Mosaicism"),
	("XL", "X-linked"),("XLD", "X-linked Dominant"), ("XLR", "X-linked Recessive"), ("DR", "Digenic Recessive"), ("ICB", "Inherited chromosomal imbalance"), ("DD", "Digenic Dominant")])
	for key,value in ph.items():
		if key in forsub:
			inttext1+=key+" - "+value+";"
	for key2, value2 in inh.items():
		if key2 in forsub:
			inttext2+=key2+"-"+value2+";"
	if len(inttext1)==1:
		text1+="."
	else:
		text1+=inttext1
	text1+=text2
	if len(inttext2)==1:
		text1+="."
	else:
		text1+=inttext2
	text1+=text3	
	ws2["A4"].value=text1


def final_things(ts, ch, bp,twobps, is_region, bb, name, is_first, is_second, params, br):
	"""Execute all the functions before. This functions organizes the process of tabel construction, 
	after the identification of TADs: gets the genes for the TADs, defines the titles for the output,
	Trites the output."""
	dd=list(ts.keys())
	forsub2=set()
	for key, value in ts.items():
		if (key=="brTAD_" or key=="TBR_") and twobps==True:
			dic1, dic_text1, forsub1=final_things_sub(ch, value, bp[0], key,"_A", bp[1], params)
			dic2, dic_text2, forsub2=final_things_sub(ch, value, bp[1], key, "_B", bp[0], params)
			ndic=merge_two_ordered_dicts(dic1, dic2)
			dic=read_snps_new.make_regions(read_snps_new.read_regions(value, ndic), ndic,ch)
			dic_text={**dic_text1, **dic_text2}
			params["dic1"]=ndic
			params["dictext1"]=dic_text
		if (key=="brTAD_" or key=="TBR_") and twobps==False:
			ndic, dic_text, forsub1=final_things_sub(ch, value, bp, key, "", "", params)
			dic=read_snps_new.make_regions(read_snps_new.read_regions(value, ndic), ndic,ch)
			if "dic1" not in params:
				params["dic1"]=ndic
				params["dictext1"]=dic_text
			else:
				params["dic2"]=ndic
				params["dictext2"]=dic_text
		if key!="brTAD_" and key!="TBR_":############################################REVER
			ndic, dic_text, forsub1=final_things_sub(ch, value, bp[0], key, "", "", params)
			dic=read_snps_new.make_regions(read_snps_new.read_regions(value, ndic), ndic,ch)
		if key=="pter_br_":
			cor="pter region at "+ch+":"+"{:,}".format(value[0])+"-"+"{:,}".format(value[1])
		if key=="qter_br_":
			cor="qter region at "+ch+":"+"{:,}".format(value[0])+"-"+"{:,}".format(value[1])
		if key=="brTAD_":
			if params["tt"]!="Spec_Rg":
				cor=ch+bb+" breakpoint within "+params["tad"][0]+" TAD at "+ch+":"+"{:,}".format(value[0])+"-"+"{:,}".format(value[1])
			else:
				cor=ch+bb+" "+params["tad"][0]+" TAD at "+ch+":"+"{:,}".format(value[0])+"-"+"{:,}".format(value[1])
		if key.startswith("TAD"):
			cor=params["tad"][0]+" "+key[:-1]+" at chr"+ch+":"+"{:,}".format(value[0])+"-"+"{:,}".format(value[1])
		if key=="TBR_":
			if params["tt"]!="Spec_Rg":
				cor=ch+bb+" breakpoint within "+params["tad"][0]+" interTAD region at chr"+ch+":"+"{:,}".format(value[0])+"-"+"{:,}".format(value[1])
			else:
				cor=ch+bb+" "+params["tad"][0]+" interTAD region at chr"+ch+":"+"{:,}".format(value[0])+"-"+"{:,}".format(value[1])
		if key.startswith("TDel"):
			if "TDel" not in params:
				params["TDel"]=dict(ndic)
				params["TDel_tx"]=dict(dic_text)
			else:
				z =dict(ndic)
				params["TDel"]={**z,  **params["TDel"]}
				z = dict(dic_text)
				params["TDel_tx"]= {**z,  **params["TDel_tx"]}
			cor=params["tad"][0]+" TAD at chr"+ch+":"+"{:,}".format(value[0])+"-"+"{:,}".format(value[1])
		if key.startswith("interTDel") or key.startswith("TBR"):
			if  key.startswith("interTDel"): 
				if "TDel" not in params:
					params["TDel"]=dict(ndic)
					params["TDel_tx"]=dict(dic_text)
				else:
					z = dict(ndic)
					params["TDel"]= {**z,  **params["TDel"]}
					z =dict(dic_text)
					params["TDel_tx"]={**z,  **params["TDel_tx"]}
			cor=params["tad"][0]+" interTAD region at chr"+ch+":"+"{:,}".format(value[0])+"-"+"{:,}".format(value[1])
		ws2.append([cor, "", "","",""])
		#forhtml.write_html("fullline", [cor, "", "","",""], "tool_temp_data/btemp.html")
		forsub=forsub1.union(forsub2)
		if len(dic)==0:
			make_report_table("None",{}, {}, ch, name,value, params, forsub, bp, br, ch+bb)
		else:
			for key2, value2 in dic.items():
				make_report_table(key2,dic, dic_text, ch, name,value, params, forsub, bp, br, ch+bb)

def merge_tads_sur(t1, t2):
	"""when analysing a deletion, this function makes
	sure that there is no TAD repeted on the intervals.
	retruns it without repetition."""
	if "brTAD_" in t1:
		aa=t1["brTAD_"][0]
	else:
		aa=t1["TBR_"][0]
	if "brTAD_" in t2:
		bb=t2["brTAD_"][0]
	else:
		bb=t2["TBR_"][0]	
	t11=t1.copy()
	t22=t2.copy()
	for key,value in t1.items():
		if value[0]>aa:
			del t11[key]
	for key2,value2 in t2.items():
		if value2[0]<bb:
			del t22[key2]
	return t11, t22

def deals_with_deletions(ch, br, is_region, b1, b2, name, params):
	"""method used by get_results to deal with deletions, 
	duplications or specific regions search."""
	intnewt1=newtads(ch[0], int(float(br[0])), params["tad"][1],  params["tad"][-2], params, 0)#gets the tads for the breakpoint1 region
	intnewt2=newtads(ch[0], int(float(br[1])), params["tad"][1],  params["tad"][-2], params, 0)#gets the tads for breakpoint2 region
	if ch[0]=="Y":
		intnewt1["brTAD_"][1]=intnewt2["brTAD_"][1]
	if intnewt1!=intnewt2 and ch[0]!="Y":#checks if the deletion/duplication/region is not all in the same tad. if not removes the TAD+1/TAD-1 that may affect the structure of the table, disallowing the repetition of TADs
		intnewt1, intnewt2=merge_tads_sur(intnewt1,intnewt2)
		tdel=get_tdel(intnewt1, intnewt2)#if its not see if theire is more than both brTAds affected by the deletion/duplication/specific region
		if tdel!={}:#if there is at least one more tad affected
			final_things(intnewt1, ch[0], br[0], False, is_region, b1, name, True, False, params, br)#runs the gene search and writing for the first breakpoint region
			newtds, params=deleted_tads(tdel["TDel"], ch[0], params)#Selects the new tads affected by the alteration
			final_things(newtds, ch[0], "0",False, is_region, b1, name, False, False, params, br)#alt#and runs the gene search and writing on them
			final_things(intnewt2,ch[0], br[1], False, is_region, b2, name, False, True, params, br)#runs the gene search for the second breakpoint
		else:
			final_things(intnewt1, ch[0], br[0], False, is_region, b1, name, True, False, params, br)#runs the gene search and writing for the first breakpoint region
			final_things(intnewt2,ch[0], br[1], False, is_region, b2, name, False, True, params, br)#runs the gene search and writing for the second breakpoint region
	if ch[0]=="Y" or intnewt1==intnewt2:
		final_things(intnewt1, ch[0], br, True, is_region, b1, name, True, True, params, br)#if both breakpoints are on the same TAD, runs the gene search and writing for the respective TAD
	global ws2
	if params["tt"]=="Deletion" or params["tt"]=="Duplication":
		ws2.append([params["fntxt"]])
		make_format_ws2(ws2, True, False, params)
		ws2=wb.create_sheet()
		ws2.title="ACMG_calculator"
		calc_ACMG.make_ACMG(params, ch, br[0], br[1], ws2)
	if params["tt"]=="Spec_Rg":
		ws2.append([params["fntxt"]])
		make_format_ws2(ws2, True, False, params)
	#forhtml.write_end(params["html_A"])

def deals_with_trans_with_dels(br, ch, newname1, newname2, b1, b2, name1, name2, params):
	"""method used by get_results to deal with 
	translocations/inversions associated with deletions"""
	br1=br[0].split("-")
	br2=br[1].split("-")
	temp_tads=params["tads"].copy()
	prepare_exel(newname1, True, ch[0]+b1, params["coments"], params)#opens a tab on exel whre the table will be written
	if len(br1)==1:#If there is no region defined on br1, it makes it normally
		intnewt1=newtads(ch[0], int(float(br1[0])), params["tad"][1],  params["tad"][-2], params, 0)#gets the tads for br1
		final_things(intnewt1, ch[0], br1[0], False, False, b1, name1, True, True, params, [])#runs the gene search and writing for the first breakpoint
	if len(br1)>1:#if there is more than one value, assumes deletion, and sends this to the deletion function
		deals_with_deletions([ch[0]], br1, False, b1, b1, name1, params)###
	params["tads"]=temp_tads
	if ch[0]!=ch[1] or params["tt"]=="Insertion":#checks if cromossome 1 is the same as chromossome 2. if it is, its an inversion, and only one tab of the exel is used. otherwise, opens another
		ws2.append([params["fntxt"]])
		make_format_ws2(ws2, True, False, params)
		#forhtml.write_end(params["html_A"])
		prepare_exel(newname2, False, ch[1]+b2, params["coments"], params)
	if len(br2)>1:#if there is more than one value, assumes deletion, and sends this to the deletion function
		deals_with_deletions([ch[1]], br2, False, b2, b2, name2, params)###
		ws2.append([params["fntxt"]])
		make_format_ws2(ws2, True, False, params)
	if len(br2)==1:#If there is no region defined on br2, it makes it normally
		intnewt2=newtads(ch[1], int(float(br2[0])),  params["tad"][1],  params["tad"][-2], params, 1)
		final_things(intnewt2, ch[1], br2[0],False, False, b2, name2, True ,True, params, [])
		ws2.append([params["fntxt"]])
		make_format_ws2(ws2, False, False, params)
	#if "html_B" in params:
		#forhtml.write_end(params["html_B"])
	#else:
		#forhtml.write_end(params["html_A"])
	#para A
	#para B

def deal_with_translocations_inversions(ch, br, newname1, newname2, b1, b2, name1, name2, params):
	"""deals with simple translocations and inversions. Used by get_results"""
	intnewt1=newtads(ch[0], int(float(br[0])), params["tad"][1],  params["tad"][-2], params, 0)#gets the tads for br1
	intnewt2=newtads(ch[1], int(float(br[1])), params["tad"][1],  params["tad"][-2], params, 1)#gets the tads for br2
	prepare_exel(newname1, True, ch[0]+b1, params["coments"], params)
	final_things(intnewt1, ch[0], br[0], False, False, b1, name1, True, True, params, [])
	if intnewt1!=intnewt2:
		#forhtml.write_end(params["html_A"])
		ws2.append([params["fntxt"]])
		make_format_ws2(ws2, False, False, params)
		if newname2=="":
			newname2=newname1
		prepare_exel(newname2, False, ch[1]+b2, params["coments"], params)
		final_things(intnewt2, ch[1], br[1], False, False, b2, name2, True, True, params,[])
	ws2.append([params["fntxt"]])
	make_format_ws2(ws2, False, False, params)


def get_results(cA, cB, brA, brB, tt, der, params):
	"""the method that deals with the information from the formula and distributes the work
	depending on the type of alteration. it also defines the nomenclature of the breakpoint"""
	name1=""
	name2=""
	ch=[cA,cB]
	br=[brA,brB]
	if tt=="Insertion":
		b1=read_cytoband(params["band"], ch[0], br[0].split("-")[0])
		b2=read_cytoband(params["band"], ch[1], br[1].split("-")[0])
		name1=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, False, False, False, False, "ins", ch[0], params["dic_ref"])
		name2=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, False, False, False, False, "ins", ch[1], params["dic_ref"])
		newname1=["Table 1 - Characterization of the insertion region "+ch[0]+b1,"","",""]
		newname2=["Table 2 - Characterization of the excision region "+ch[1]+b2,"","",""]
		deals_with_trans_with_dels(br, ch, newname1, newname2, b1, b2, name1, name2, params)
	elif tt=="Deletion":
		b1=read_cytoband(params["band"], ch[0], br[0])
		b2=read_cytoband(params["band"], ch[0], br[1])
		name1=pq_nomenclature(ch[0], ch[0], b1, b2, br[0], br[1], False, False, False, False, False, "del", "", params["dic_ref"])
		newname=["Table 1 - Characterization of the deleted region "+"del("+ch[0]+")("+b1+";"+b2+")","","",""]
		prepare_exel(newname, True, ch[0]+b1, params["coments"], params)
		deals_with_deletions(ch, br, False, b1,b2, name1, params)
	elif tt=="Spec_Rg":
		b1=read_cytoband(params["band"], ch[0], br[0])
		b2=read_cytoband(params["band"], ch[0], br[1])
		name1="Query Region - chr"+ch[0]+":"+str(br[0])+"-"+str(br[1])
		newname=["Table 1 - Characterization of chosen region "+ch[0]+b1+b2,"","",""]
		prepare_exel(newname, True, ch[0]+b1, params["coments"], params)
		deals_with_deletions(ch, br, False, b1,b2, name1, params)
	elif tt=="Duplication":
		b1=read_cytoband(params["band"], ch[0], br[0])
		b2=read_cytoband(params["band"], ch[0], br[1])
		name1=pq_nomenclature(ch[0], ch[0], b1, b2, br[0], br[1], False, False, False, False, False, "dup", "", params["dic_ref"])
		newname=["Table 1 - Characterization of the duplicated region "+"dup("+ch[0]+")("+b1+";"+b2+")","","",""]
		prepare_exel(newname, True, ch[0]+b1, params["coments"], params)
		deals_with_deletions(ch, br, False, b1,b2, name1, params)
	else:
		if len(br[0].split("-"))>1:
			if len(br[0].split("-"))==2:
				if int(br[0].split("-")[1])-int(br[0].split("-")[0])<=1000:
					tt2=int(br[0].split("-")[1])-int(br[0].split("-")[0])
					brA=str(int(br[0].split("-")[0])+tt2/2)
					v1=False
					v2=False
				else:
					brA=br[0]
					v1=True
					v2=False
			if len(br[0].split("-"))==3:
				brA=br[0].split("-",1)[1]
				v1=False
				v2=True
		if len(br[0].split("-"))==1:
			brA=br[0]
			v1=False
			v2=False
		if len(br[1].split("-"))==1:
			brB=br[1]
			v3=False
			v4=False		
		if len(br[1].split("-"))>1:
			if len(br[1].split("-"))==2:
				if int(br[1].split("-")[1])-int(br[1].split("-")[0])<=1000:
					tt2=int(br[1].split("-")[1])-int(br[1].split("-")[0])
					brB=str(int(br[1].split("-")[0])+tt2/2)
					v3=False
					v4=False
				else:
					brB=br[1]
					v3=True
					v4=False
			if len(br[1].split("-"))==3:
				brB=br[1].split("-",1)[1]
				v3=False
				v4=True
		br=[brA,brB]
		if len(br[0].split("-"))>1 or len(br[1].split("-"))>1 or tt=="Unbalanced_translocation":
			b1=read_cytoband(params["band"], ch[0], br[0].split("-")[0])
			b2=read_cytoband(params["band"], ch[1], br[1].split("-")[0])
			if tt=="Balanced_translocation":
				newname1=["Table 1 - Characterization of the breakpoint region "+cA+b1.split("-")[0]+" of t("+cA+";"+cB+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				newname2=["Table 2 - Characterization of the breakpoint region "+cB+b2.split("-")[0]+" of t("+cA+";"+cB+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				name1=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, v1, v2, v3, v4, "trans", ch[0], params["dic_ref"])
				name2=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, v1, v2, v3, v4, "trans", ch[1], params["dic_ref"])
			elif tt=="Inversion":
				newname1=["Table 1 - Characterization of inv("+cA+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				newname2=""
				name1=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, v1, v2, v3, v4, "inv", ch[0], params["dic_ref"])
				name2=name1
			elif tt=="Unbalanced_translocation":
				newname1=["Table 1 - Characterization of the breakpoint region "+cA+b1.split("-")[0]+" of the unbalanced t("+cA+";"+cB+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				newname2=["Table 2 - Characterization of the breakpoint region "+cB+b2.split("-")[0]+" of the unbalanced t("+cA+";"+cB+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				sz=read_chr_size(params["chrsz"])
				if der==ch[0]:
					name1=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, v1, v2, v3, v4, "trans", ch[0], params["dic_ref"])
					if "p" in b1 and "p" in b2:
						name2=pq_nomenclature(ch[1], ch[1], b1, b2, "1",br[1], False, False, False, False, False, "dup", "", params["dic_ref"])
						br=["1-"+brA,"1-"+brB]
					if "p" in b1 and "q" in b2:
						name2=pq_nomenclature(ch[1], ch[1], b1, b2, br[1], sz[ch[1]], False, False, False, False, False, "dup", "", params["dic_ref"])
						br=["1-"+brA, brB+"-"+sz[ch[1]]]
					if "q" in b1 and "q" in b2:
						name2=pq_nomenclature(ch[1], ch[1], b1, b2, br[1], sz[ch[1]], False, False, False, False, False, "dup", "", params["dic_ref"])
						br=[brA+"-"+sz[ch[0]], brB+"-"+sz[ch[1]]]
					if "q" in b1 and "p" in b2:
						name2=pq_nomenclature(ch[1], ch[1], b1, b2, "1",br[1], False, False, False, False, False, "dup", "", params["dic_ref"])
						br=[brA+"-"+sz[ch[0]], "1-"+brB]
				if der==ch[1]:
					name2=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, v1, v2, v3, v4, "trans", ch[1], params["dic_ref"])
					if "p" in b1 and "p" in b2:
						name1=pq_nomenclature(ch[0], ch[0], b1, b2, "1",br[0], False, False, False, False, False, "dup", "", params["dic_ref"])
						br=["1-"+brA, "1-"+brB]
					if "p" in b1 and "q" in b2:
						name1=pq_nomenclature(ch[0], ch[0], b1, b2, "1",br[0], False, False, False, False, False, "dup", "", params["dic_ref"])
						br=["1-"+brA, brB+"-"+sz[ch[1]]]
					if "q" in b1 and "q" in b2:
						name1=pq_nomenclature(ch[0], ch[0], b1, b2, br[0], sz[ch[0]], False, False, False, False, False, "dup", "", params["dic_ref"])
						br=[brA+"-"+sz[ch[0]], brB+"-"+sz[ch[1]]]
					if "q" in b1 and "p" in b2:
						name1=pq_nomenclature(ch[0], ch[0], b1, b2, br[0], sz[ch[0]], False, False, False, False, False, "dup", "", params["dic_ref"])
						br=[brA+"-"+sz[ch[0]], "1-"+brB]
			deals_with_trans_with_dels(br, ch, newname1, newname2, b1, b2, name1, name2, params)
		else:
			b1=read_cytoband(params["band"], ch[0], br[0])
			b2=read_cytoband(params["band"], ch[1], br[1])
			if tt=="Balanced_translocation":
				newname1=["Table 1 - Characterization of the breakpoint region "+cA+b1.split("-")[0]+" of t("+cA+";"+cB+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				newname2=["Table 2 - Characterization of the breakpoint region "+cB+b2.split("-")[0]+" of t("+cA+";"+cB+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				name1=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, False, False, False, False, "trans", ch[0], params["dic_ref"])
				name2=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, False, False, False, False, "trans", ch[1], params["dic_ref"])
			elif tt=="Inversion":
				newname1=["Table 1 - Characterization of inv("+cA+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				newname2=""
				name1=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, False, False, False, False, "inv", ch[0], params["dic_ref"])
				name2=name1
			deal_with_translocations_inversions(ch, br, newname1, newname2, b1,b2, name1, name2, params)
	return name1, name2

def edit_link(link, outfiles, name_3):
	n=outfiles.split("/")[-1]
	f=open(link)
	lines=f.readlines()
	f.close()
	out=open(link, "w")
	for el in lines:
		if 'content="10"' not in el and 'SVInterpreter is creating you' not in el:
			out.write(el)
		elif 'SVInterpreter is creating you' in el:
			gg=el.replace('<h3><center><p style="color:red"><b>SVInterpreter is creating your table. The running time depends on the complexity of the table and the size of the region. Please be patient....</h3></center></p></b>', '')
			bb=gg.replace("</div>","")
			out.write(bb)
			out.write('<h3><center><p><a href=/outputs/'+n+'>Download report table!</center></a></p></h3>')
			if name_3 !="":
				out.write("<aaa><center><a href=/outputs/"+name_3+">Download CNV table!</a></center></aaa>")
			out.write("</div>")
	out.close()

def load_obj(gg):
	name=gg[:-5]
	with open("/var/www/html/dgrctools.insa.min-saude.pt/public_html/outputs/"+name + '.pkl', 'rb') as f:
		return pickle.load(f)

##########################RUN THE SCRIPT############################################################

def mainn(outfiles, cA, cB, brA, brB, tt, der, link, name_3):
	params=load_obj(link.split("/")[-1])
	fntxt="Analysis performed with bioinformatics tool TAD-GCtool v2.0 financiated by the FCT project HMSP-ICT/0016/2013. Human genome version "+params["version"]+". TADs according to "+params["tad"][2]+". NA - not available; ND - not determined." 
	params["fntxt"]=fntxt
	name1, name2=get_results(cA, cB, brA, brB, tt, der, params)
	wb.save(filename=outfiles)
	edit_link(link, outfiles, name_3)

mainn(argv[1], argv[2], argv[3], argv[4], argv[5], argv[6], argv[7], argv[8], argv[9])
