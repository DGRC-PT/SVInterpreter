import cross_deletions_with_dgv_results_Vfor_new
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, Fill, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side


def make_reg_to_test(params, chrr, start,end):
	dic={"ovl":"full", "version": params["version"], "perc":"70", "tt":params["tt"], "chrA":chrr[0], "brA":start+"-"+end, "dats":["gnomad", "chaisson", "collins", "DGV", "1000Genomes", "ClinGenben", "ClinGenlben"]}
	gg=cross_deletions_with_dgv_results_Vfor_new.exect_for_ACMG(dic)
	return gg

def merge_two_dicts(x, y, w):
    #z = x.copy()   # start with x's keys and values
    #z.update(y)    # modifies z with y's keys and values & returns None
	z = {**x, **y, **w}
	return z

def correct_oes(oes1):
	oes=set()
	for key,value in oes1.items():
		if "b" in value[0] or "b" in value[1]:
			oes.add(key)
	return oes

def make_ACMG(params, chrr, start, end, ws2):
	scores=OrderedDict()
	#genes afetados
	br=False
	to_count=set()
	oes=correct_oes(params["oes"])
	if "dic1" in params:
		dic1=params["dic1"]
	else:
		dic1={}
	if "dic2" in params:
		dic2=params["dic2"]
	else:
		dic2={}
	if "dictext1" in params:
		dictext1=params["dictext1"]
	else:
		dictext1={}
	if "dictext2" in params:
		dictext2=params["dictext2"]
	else:
		dictext2={}
	for el in dic1:
		if "Breakpoint" in el and br==False:
			br==True
		elif "Breakpoint" not in el and br==True:
			to_count.add(el)
		elif "Breakpoint" in el and br==True:
			break
	if len(dic2)>0:
		for ee in dic2:
			if "Breakpoint" not in ee and br==True:
				to_count.add(ee)
			elif "Breakpoint" in el and br==True:
				break
	if "TDel" in params:
		for ele in params["TDel"]:
			to_count.add(ele)
	for el in to_count:
		if el in oes:
			scores["1A"]=["Genes affected by the CNV", 0]
			scores["2A"]=["Complete overlap of an HI/TS gene", 1]
	for eee in dictext2:
		to_count.add(eee)
	for ele in dictext1:
		to_count.add(ele)
	if "TDel_tx" in params:
		for tt in params["TDel_tx"]:
			to_count.add(tt)
	if len(to_count)==0:
		scores["1B"]=["No genes affected by the CNV", -0.6]
	else:
		scores["1A"]=["Genes affected by the CNV", 0]
	#os genes affectados sao haploins
	print("oes",oes)
	if "2A" not in scores:
		print("dic", dictext1, dictext2)
		if "TDel_tx" in params:
			z=merge_two_dicts(dictext1, dictext2,params["TDel_tx"])
		else:
			z=merge_two_dicts(dictext1, dictext2,{})
		print("z",z)
		if len(z)>0:
			for w,value in z.items():
				if w in oes:
					print(w,value)
					if params["tt"]=="Duplication":
						if "2I-2L" not in scores:
							scores["2I-2L"]=["One or both breakpoints affect an HI/TS gene", "0 to 0.9"]
					elif "5'" in value[0] and "2C1" not in scores:
						scores["2C1"]=["Affects the 5' UTR of the gene and codding sequence", 0.9]
					elif "3'" in value[0] and "2D2-2D4" not in scores:
						scores["2D2-2D4"]=["Affects the 3' UTR of the gene and coding sequence", "0.3 to 0.9"]
					else:
						scores["2E"]=["Both breakpoints are located in the same gene", "0 to 0.9"]
	#make the overlap
	gg=make_reg_to_test(params, chrr, start, end)
	if True in gg and params["tt"]=="Deletion":
		scores["2F"]=["Completely contained within an established benign CNV region", -1]
	elif True in gg and params["tt"]=="Duplication":
		scores["2C,2D or 2F"]=["A benign duplication having the same gene content as the input", "-0.9 to -1"]
	if len(to_count)<=34:
		scores["3A"]=["Number of affected genes: 0-34",0]
	elif len(to_count)>34 and len(to_count)<=49:
		scores["3B"]=["Number of affected genes: 35-49", 0.45]
	else:
		scores["3C"]=["Number of affected genes: >49", 0.9]
	print(scores)
	if params["tt"]=="Deletion":
		write_ACMG_del(scores, ws2)
	else:
		write_ACMG_dup(scores, ws2)
	

def write_ACMG_del(scores, ws2):
	dell={"1A":["7"], "1B":["8"], "2A":["10"], "2C1":["13"], "2D2-2D4":["17","18","19"], "2E":["20"], "2F":["21"], "3A":["25"], "3B":["26"], "3C":["27"]}
	ws2.append(["CNV Interpretation Scoring Rubric: Copy Number LOSS"])
	ws2.append(["The parameters presented bellow are according to ", '=HYPERLINK("https://cnvcalc.clinicalgenome.org/cnvcalc/", "ClinGen CNV Pathogenicity Calculator")'])
	ws2.append(['Parameters with grey background were automatically filled. Parameters 2D-2, 2D-3 and 2D-4 must be reviewed manually.'])
	ws2.append([''])
	ws2.append(["", "Sujested Points", "Maximum Score", "Points Given"])
	ws2.append(["Section 1: Initial Assessment of Genomic Content"])
	ws2.append(["1A. Contains protein-coding or other known functionally important elements", "0", "-0.6"])
	ws2.append(["1B. Does NOT contain protein-coding or any known functionally important elements", "0", "-0.6"])
	ws2.append(["Section 2 : Overlap with Established/Predicted HI or Established Benign Genes/Genomic Regions"])
	ws2.append(["2A. Complete overlap of an established HI gene/genomic region", "1", "1"])
	ws2.append(["2B. Partial overlap of an established HI genomic region", "0", "0"])
	ws2.append(["2C. Partial overlap with the 5’ end of an established HI gene (3’ end of the gene not involved)..."])
	ws2.append(["2C-1. …and coding sequence is involved", "0.90 (Range : 0.45 to 1.00)", "1"])
	ws2.append(["2C-2. …and only the 5’ UTR is involved", "0 (Range : 0 to 0.45)", "0.45"])
	ws2.append(["2D. Partial overlap with the 3’ end of an established HI gene (5’ end of the gene not involved)…"])
	ws2.append(["2D-1. …and only the 3’ untranslated region is involved.", "0", "0"])
	ws2.append(["2D-2. …and only the last exon is involved. Other established pathogenic variants have been reported in this exon.", "0.90 (Range : 0.45 to 0.90)", "0.9"])
	ws2.append(["2D-3. …and only the last exon is involved. No other established pathogenic variants have been reported in this exon.", "0.30 (Range : 0 to 0.45)", "0.45"])
	ws2.append(["2D-4. …and it includes other exons in addition to the last exon. Nonsense-mediated decay is expected to occur.", "0.90 (Range : 0.45 to 1.00)", "1"])
	ws2.append(["2E. Both breakpoints are within the same gene (gene-level sequence variant)", "(Range : 0 to 0.90)", ""])
	ws2.append(["2F. Completely contained within an established benign CNV region"])
	ws2.append(["2G. Overlaps an established benign CNV, but includes additional genomic material", "0", "0"])
	ws2.append(["2H. Multiple HI predictors suggest that AT LEAST ONE gene in the interval is haploinsufficient (HI)", "0.15", "0.15"])
	ws2.append(["Section 3: Evaluation of Gene Number"])
	ws2.append(["3A. 0-24 Genes", "0", "0"])
	ws2.append(["3B 25-34 Genes", "0.45", "0.45"])
	ws2.append(["3C. 35 or more Genes", "0.9", "0.9"])
	ws2.append(["Section 4: Detailed Evaluation of Genomic Content Using Published Literature, Public Databases, and/or Internal Lab Data"])
	ws2.append(["Reported proband has either:"])
	ws2.append(["A complete deletion of or a LOF variant within gene encompassed by the observed copy number loss OR"])	
	ws2.append(["an overlapping copy number loss similar in genomic content to the observed copy number loss AND…"])	
	ws2.append(["4A. …the reported phenotype is highly specific and relatively unique to the gene or genomic region", "(Range : 0.15 to 0.45)", "0.9 (total)"])	
	ws2.append(["4B. …the reported phenotype is consistent with the gene/genomic region, is highly specific, but not necessarily unique to the gene/genomic region", "(Range : 0.15 to 0.45)", "0.9 (total)"])	
	ws2.append(["4C. …the reported phenotype is consistent with the gene/genomic region, but not highly specific and/or with high genetic heterogeneity", "(Range : 0 to 0.30)", "0.9 (total)"])	
	ws2.append(["4D.…the reported phenotype is NOT consistent with what is expected for the gene/genomic region or not consistent in general", "0 (Range: 0 to -0.30)", "-0.3 (total)"])	
	ws2.append(["4E. Reported proband has a highly specific phenotype consistent with the gene/genomic region, but the inheritance of the variant is unknown.", "0.1 (Range : 0 to 0.15)", "0.3 (total)"])	
	ws2.append(["4F. 3-4 observed segregation", "0.15", "0.15"])	
	ws2.append(["4G. 5-6 observed segregation", "0.3", "0.3"])	
	ws2.append(["4H. 7 or more observed segregation", "0.45", "0.45"])	
	ws2.append(["4I. Variant is NOT found in another individual in the proband’s family AFFECTED with a consistent, specific, well-defined phenotype (no known phenocopies)", "-0.45 (Range: 0 to -0.45)", "-0.45"])	
	ws2.append(["4J. Variant IS found in another individual in the proband’s family UNAFFECTED with the specific, well-defined phenotype observed in the proband", "-0.3 (Range: 0 to -0.30)", "-0.3"])	
	ws2.append(["4K. Variant IS found in another individual in the proband’s family UNAFFECTED with the non-specific phenotype observed in the proband", "-0.15-0.3(Range: 0 to -0.15)", "-0.3"])		
	ws2.append(["4L. Statistically significant increase amongst observations in cases (with a consistent, specific, well-defined phenotype) compared to controls", "0.45 (range: 0 to 0.45)", "0.45"])		
	ws2.append(["4M. Statistically significant increase amongst observations in cases (without a consistent, non-specific phenotype OR unknown phenotype) compared to controls", "0.30 (range: 0 to 0.30)", "0.45"])		
	ws2.append(["4N. No statistically significant difference between observations in cases and controls", "-0.9 (Range :0 to -0.9)", "-0.9"])		
	ws2.append(["4O. Overlap with common population variation", " -1.00 (Range :0 to -1.00)", "-1"])		
	ws2.append(["Section 5: Evaluation of Inheritance Pattern/Family History for Patient Being Studied"])		
	ws2.append(["5A. Use appropriate category from de novo scoring section in Section 4.", "", "0.45"])		
	ws2.append(["5B. Patient with specific, well-defined phenotype and no family history. CNV is inherited from an apparently unaffected parent.", "-0.30 (Range : 0 to -0.45)", "-0.45"])
	ws2.append(["5C. Patient with non-specific phenotype and no family history. CNV is inherited from an apparently unaffected parent.", "-0.15 (Range : 0 to -0.30)", "-0.3"])		
	ws2.append(["5D. CNV segregates with a consistent phenotype observed in the patient’s family. (use scoring from section 4)", "", "0.45"])	
	ws2.append(["5E. Use appropriate category from non-segregation section in Section 4. (use scoring from section 4)", "", "-0.45"])	
	ws2.append(["5F. Inheritance information is unavailable or uninformative", "0", "0"])	
	ws2.append(["5G. Inheritance information is unavailable or uninformative. The patient phenotype is non-specific, but is consistent with what has been described in similar cases.", "0.10 (Range : 0 to 0.15)", "0.15"])	
	ws2.append(["5H. Inheritance information is unavailable or uninformative. The patient phenotype is highly specific and consistent with what has been described in similar cases.", "0.30 (Range : 0 to 0.30)", "0.30"])	
	ws2.append([""])
	ws2.append(["", "", "Total","=SUM(A7:D55)"])
	ws2.append(["","","Classification",'=IF(D57>=0.99,"Pathogenic",IF(AND(D57>=0.9,D57<=0.98),"Potentialy pathogenic",IF(AND(D57>=-0.89,D57<=0.89),"VUS",IF(AND(D57>=-0.98,D57<=-0.9),"Potentialy benign","Benign"))))'])
	ll=["A","B","C", "D"]
	for key,value in scores.items():
		for el in dell[key]:
			gg=ws2["D"+el]
			gg.value=value[-1]
			for ele in ll:
				ws2[ele+el].fill=PatternFill("solid", fgColor="DDDDDD")
	make_ACMG_table_format(ws2)



def make_ACMG_table_format(ws2):
	didths={"A":101.86, "B":21,"C":18.86,"D":11.14}
	for key,value in didths.items():
		ws2.column_dimensions[key].width = value
	i=6
	cols=["A", "B", "C", "D"]
	ws2["A1"].font=Font(size=18, name="Arial Narrow")
	ws2["A2"].font=Font(size=10, name="Arial Narrow")
	ws2["B2"].font=Font(color="0000CC", size=10, name="Arial Narrow")
	ws2["A3"].font=Font(size=10, name="Arial Narrow")
	ws2["B5"].font=Font(size=10, name="Arial Narrow", bold=True)
	ws2["C5"].font=Font(size=10, name="Arial Narrow", bold=True)
	ws2["D5"].font=Font(size=10, name="Arial Narrow", bold=True)
	while i<=ws2.max_row:
		index="A"+str(i)
		bb=ws2[index].value
		if bb!=None:
			if bb.startswith("Section"):
				ws2[index].font=Font(name="Arial Narrow", size=10, bold=True, color="009900")
				for el in cols:
					ws2[el+str(i)].border=Border(bottom=Side(style='thin'))
			elif ws2["C"+str(i)].value=="Total":
				ws2["C"+str(i)].font=Font(size=10, name="Arial Narrow", bold=True)
				ws2["D"+str(i)].font=Font(size=10, name="Arial Narrow", bold=True)
				ws2["C"+str(i)].border=Border(bottom=Side(style='thin'))
				ws2["D"+str(i)].border=Border(bottom=Side(style='thin'))
				ws2["C"+str(i+1)].font=Font(size=10, name="Arial Narrow", bold=True)
				ws2["D"+str(i+1)].font=Font(size=10, name="Arial Narrow", bold=True)
				ws2["C"+str(i+1)].border=Border(bottom=Side(style='thin'))
				ws2["D"+str(i+1)].border=Border(bottom=Side(style='thin'))
			elif bb!="":
				for el in cols:
					ws2[el+str(i)].font=Font(size=10, name="Arial Narrow")
		i+=1


def write_ACMG_dup(scores, ws2):
	dell={"1A":["7"], "1B":["8"], "2A":["10"], "2I-2L":["18", "19","20","21"], "2C,2D or 2F":["12","13","15"], "3A":["23"], "3B":["24"], "3C":["25"]}
	ws2.append(["CNV Interpretation Scoring Rubric: Copy Number GAIN"])
	ws2.append(["The parameters presented bellow are according to ", '=HYPERLINK("https://cnvcalc.clinicalgenome.org/cnvcalc/", "ClinGen CNV Pathogenicity Calculator")'])
	ws2.append(['Parameters with grey background were automatically filled.'])#####
	ws2.append([''])
	ws2.append(["", "Sujested Points", "Maximum Score", "Points Given"])
	ws2.append(["Section 1: Initial Assessment of Genomic Content"])
	ws2.append(["1A. Contains protein-coding or other known functionally important elements", "0", "0"])
	ws2.append(["1B. Does NOT contain protein-coding or any known functionally important elements", "-0.6", "-0.6"])
	ws2.append(["Section 2: Overlap with Established Triplosensitive (TS), Haploinsufficient (HI), or Benign Genes or Genomic Regions"])
	ws2.append(["2A. Complete overlap; the TS gene or minimal critical region is fully contained within the observed copy number gain", "1", "1"])
	ws2.append(["2B. Partial overlap of an established TS genomic region", "0", "0"])
	ws2.append(["2C. Identical in gene content to the established benign copy number gain", "-1","-1"])
	ws2.append(["2D. Smaller than established benign copy number gain, breakpoint(s) does not interrupt protein-coding genes", "-1", "-1"])
	ws2.append(["2E. Smaller than established benign copy number gain, breakpoint(s) potentially interrupts protein-coding gene", "0", "0"])
	ws2.append(["2F. Larger than known benign copy number gain, does not include additional protein-coding genes", "-0.90 (Range: 0 to -1.00)", "-1"])
	ws2.append(["2G. Overlaps a benign copy number gain but includes additional genomic material", "0", "0"])
	ws2.append(["2H. HI gene fully contained within observed copy number gain", "0", "0"])
	ws2.append(["2I. Both breakpoints are within the same gene (gene-level sequence variant, possibly resulting in loss of function (LOF))", "(Range : 0 to 0.9)", "0.9"])
	ws2.append(["2J. One breakpoint is within an established HI gene, patient’s phenotype is either inconsistent with what is expected for LOF of that gene OR unknown", "0", "0"])
	ws2.append(["2K. One breakpoint is within an established HI gene, patient’s phenotype is highly specific and consistent with what is expected for LOF of that gene", "0.45", "0.45"])
	ws2.append(["2L. One or both breakpoints are within gene(s) of no established clinical significance", "0","0"])
	ws2.append(["Section 3: Evaluation of Gene Number"])
	ws2.append(["3A. 0-24 Genes", "0", "0"])
	ws2.append(["3B 25-34 Genes", "0.45", "0.45"])
	ws2.append(["3C. 35 or more Genes", "0.9", "0.9"])
	ws2.append(["Section 4: Detailed Evaluation of Genomic Content Using Published Literature, Public Databases, and/or Internal Lab Data"])
	ws2.append(["Reported proband has either:"])
	ws2.append(["complete duplication of one or more genes within the observed copy number gain OR"])	
	ws2.append(["an overlapping copy number gain similar in genomic content to the observed copy number gain AND…"])	
	ws2.append(["4A. …the reported phenotype is highly specific and relatively unique to the gene or genomic region", "(Range : 0.15 to 0.45)", "0.9 (total)"])	
	ws2.append(["4B. …the reported phenotype is consistent with the gene/genomic region, is highly specific, but not necessarily unique to the gene/genomic region", "(Range : 0.15 to 0.45)", "0.9 (total)"])	
	ws2.append(["4C. …the reported phenotype is consistent with the gene/genomic region, but not highly specific and/or with high genetic heterogeneity", "(Range : 0 to 0.30)", "0.9 (total)"])	
	ws2.append(["4D.…the reported phenotype is NOT consistent with what is expected for the gene/genomic region or not consistent in general", "0 (Range: 0 to -0.30)", "-0.3 (total)"])	
	ws2.append(["4E. Reported proband has a highly specific phenotype consistent with the gene/genomic region, but the inheritance of the variant is unknown.", "0.1 (Range : 0 to 0.15)", "0.3 (total)"])	
	ws2.append(["4F. 3-4 observed segregation", "0.15", "0.15"])	
	ws2.append(["4G. 5-6 observed segregation", "0.3", "0.3"])	
	ws2.append(["4H. 7 or more observed segregation", "0.45", "0.45"])	
	ws2.append(["4I. Variant is NOT found in another individual in the proband’s family AFFECTED with a consistent, specific, well-defined phenotype (no known phenocopies)", "-0.45 (Range: 0 to -0.45)", "-0.45"])	
	ws2.append(["4J. Variant IS found in another individual in the proband’s family UNAFFECTED with the specific, well-defined phenotype observed in the proband", "-0.3 (Range: 0 to -0.30)", "-0.3"])	
	ws2.append(["4K. Variant IS found in another individual in the proband’s family UNAFFECTED with the non-specific phenotype observed in the proband", "-0.15-0.3(Range: 0 to -0.15)", "-0.3"])		
	ws2.append(["4L. Statistically significant increase amongst observations in cases (with a consistent, specific, well-defined phenotype) compared to controls", "0.45 (range: 0 to 0.45)", "0.45"])		
	ws2.append(["4M. Statistically significant increase amongst observations in cases (without a consistent, non-specific phenotype OR unknown phenotype) compared to controls", "0.30 (range: 0 to 0.30)", "0.45"])		
	ws2.append(["4N. No statistically significant difference between observations in cases and controls", "-0.9 (Range :0 to -0.9)", "-0.9"])		
	ws2.append(["4O. Overlap with common population variation", " -1.00 (Range :0 to -1.00)", "-1"])		
	ws2.append(["Section 5: Evaluation of Inheritance Pattern/Family History for Patient Being Studied"])		
	ws2.append(["5A. Use appropriate category from de novo scoring section in Section 4.", "", "0.45"])		
	ws2.append(["5B. Patient with specific, well-defined phenotype and no family history. CNV is inherited from an apparently unaffected parent.", "-0.30 (Range : 0 to -0.45)", "-0.45"])
	ws2.append(["5C. Patient with non-specific phenotype and no family history. CNV is inherited from an apparently unaffected parent.", "-0.15 (Range : 0 to -0.30)", "-0.3"])		
	ws2.append(["5D. CNV segregates with a consistent phenotype observed in the patient’s family. (use scoring from section 4)", "", "0.45"])	
	ws2.append(["5E. Use appropriate category from non-segregation section in Section 4. (use scoring from section 4)", "", "-0.45"])	
	ws2.append(["5F. Inheritance information is unavailable or uninformative", "0", "0"])	
	ws2.append(["5G. Inheritance information is unavailable or uninformative. The patient phenotype is non-specific, but is consistent with what has been described in similar cases.", "0.10 (Range : 0 to 0.15)", "0.15"])	
	ws2.append(["5H. Inheritance information is unavailable or uninformative. The patient phenotype is highly specific and consistent with what has been described in similar cases.", "0.30 (Range : 0 to 0.30)", "0.30"])	
	ws2.append([""])
	ws2.append(["", "", "Total","=SUM(A7:D53)"])
	ws2.append(["","","Classification",'=IF(D55>=0.99,"Pathogenic",IF(AND(D55>=0.9,D55<=0.98),"Potentialy pathogenic",IF(AND(D55>=-0.89,D55<=0.89),"VUS",IF(AND(D55>=-0.98,D55<=-0.9),"Potentialy benign","Benign"))))'])
	ll=["A","B","C", "D"]
	for key,value in scores.items():
		for el in dell[key]:
			gg=ws2["D"+el]
			gg.value=value[-1]
			for ele in ll:
				ws2[ele+el].fill=PatternFill("solid", fgColor="DDDDDD")
	make_ACMG_table_format(ws2)
